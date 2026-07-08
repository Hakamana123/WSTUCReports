"""Excel export of the per-student tracking report."""

from datetime import date
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from student_tracker.segmentation import segment_counts, ALL_SEGMENTS

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
SEGMENT_FILLS = {
    "S1 - Never engaged":          PatternFill("solid", fgColor="C00000"),
    "S2 - Pre-block ghost":        PatternFill("solid", fgColor="E26B0A"),
    "S3 - W1 ghost":               PatternFill("solid", fgColor="FFC000"),
    "S4 - Dropped this week":      PatternFill("solid", fgColor="FFD966"),
    "S5 - Returning engager":      PatternFill("solid", fgColor="A9D08E"),
    "S6 - Fading engager":         PatternFill("solid", fgColor="F4B084"),
    "S7 - True sustainer":         PatternFill("solid", fgColor="70AD47"),
    "S8 - Long-tail dropout":      PatternFill("solid", fgColor="C65911"),
    "Active (single week of data)": PatternFill("solid", fgColor="BDD7EE"),
    "Unclassified":                PatternFill("solid", fgColor="D9D9D9"),
}


def _write_df(ws, df: pd.DataFrame, start_row: int = 1) -> None:
    """Write a DataFrame to a worksheet starting at start_row, with header styling."""
    df = df.copy()
    df = df.astype(object).where(df.notna(), None)
    rows = list(dataframe_to_rows(df, index=False, header=True))
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="left", vertical="center")
    for c_idx in range(1, len(df.columns) + 1):
        col_vals = ["" if v is None else str(v) for v in df.iloc[:, c_idx - 1].tolist()[:200]]
        max_len = max([len(str(df.columns[c_idx - 1]))] + [len(v) for v in col_vals])
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 2, 40)


def build_workbook(
    classified: pd.DataFrame,
    weeks: list[int],
    block_start_date: date,
    subject_code: str | None,
    reference_date: date,
    fade_threshold: float,
) -> Workbook:
    """Build the Excel workbook for download."""
    wb = Workbook()

    teaching_weeks = [w for w in weeks if w >= 1]

    # === Summary sheet ===
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Student Engagement Tracking Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Subject: {subject_code or 'Unknown'}"
    ws["A3"] = f"Block start: {block_start_date.isoformat()}"
    ws["A4"] = f"Reference date: {reference_date.isoformat()}"
    ws["A5"] = f"Fade threshold: {fade_threshold} (this week < threshold x last week on dominant signal)"
    week_list_str = ", ".join(f"W{w}" for w in weeks) + (" (W0 = pre-teaching baseline)" if 0 in weeks else "")
    ws["A6"] = f"Weeks in data: {week_list_str}"
    ws["A7"] = f"Total enrolled students: {len(classified)}"
    ws["A8"] = (
        "Segmentation driven by hours-in-subject OR login count (Option 1 — "
        "'active' = either signal > 0; fade/sustain measured on whichever "
        "signal was larger last week). Forum activity is context only."
    )

    counts = segment_counts(classified)
    _write_df(ws, counts, start_row=11)
    for row_idx, seg_name in enumerate(counts["segment"], start=12):
        fill = SEGMENT_FILLS.get(seg_name)
        if fill:
            ws.cell(row=row_idx, column=1).fill = fill

    # === Column ordering shared by Per Student + per-segment sheets ===
    front_cols = [
        "segment", "student_code", "preferred_name", "first_name", "last_name",
        "attend_type", "course",
        "last_active_week", "weeks_since_last_active", "weeks_active",
        "total_logins", "last_login_date", "days_since_last_login",
        "total_hours", "total_period_logins", "total_forum_interactions",
        "assessments_submitted", "assessments_total", "submission_rate", "avg_score_pct",
        "email_address",
    ]
    week_cols = []
    for wk in weeks:
        for suffix in ("hours", "logins", "forum"):
            c = f"W{wk}_{suffix}"
            if c in classified.columns:
                week_cols.append(c)
    other_cols = [c for c in classified.columns if c not in front_cols and c not in week_cols]
    ordered = [c for c in front_cols if c in classified.columns] + week_cols + other_cols

    def _prep(df: pd.DataFrame) -> pd.DataFrame:
        out_df = df[[c for c in ordered if c in df.columns]].copy()
        if "last_login_date" in out_df.columns:
            out_df["last_login_date"] = pd.to_datetime(
                out_df["last_login_date"], errors="coerce"
            ).dt.strftime("%Y-%m-%d").fillna("")
        if "submission_rate" in out_df.columns:
            out_df["submission_rate"] = (out_df["submission_rate"] * 100).round(1)
            out_df = out_df.rename(columns={"submission_rate": "submission_rate_pct"})
        return out_df

    # === Per-student sheet ===
    ws2 = wb.create_sheet("Per Student")
    out_df = _prep(classified)
    _write_df(ws2, out_df, start_row=1)
    seg_col_idx = list(out_df.columns).index("segment") + 1
    for r_idx, s in enumerate(out_df["segment"], start=2):
        fill = SEGMENT_FILLS.get(s)
        if fill:
            ws2.cell(row=r_idx, column=seg_col_idx).fill = fill

    # === Per-segment sheets ===
    for seg_name in ALL_SEGMENTS:
        sub = classified[classified["segment"] == seg_name]
        if sub.empty:
            continue
        sheet_name = seg_name.split(" - ")[0][:31]
        if sheet_name in [s.title for s in wb.worksheets]:
            sheet_name = sheet_name + "_"
        ws_s = wb.create_sheet(sheet_name)
        _write_df(ws_s, _prep(sub), start_row=1)

    return wb


def to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
