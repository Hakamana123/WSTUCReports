"""
Engagement Pulse Check
======================
Quick engagement segmentation using a class list and Blackboard
Performance Dashboard PDF. No usage or login report files required.

Segments students into S1–S4 + Active based on last-access date
relative to the block start date.

Dependencies:
    pip install streamlit openpyxl xlrd pdfplumber

Run with:
    streamlit run engagement_pulse_check.py
"""

import io
import re
from datetime import date, datetime, timedelta

import pdfplumber
import streamlit as st
import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ===========================================================================
# CONSTANTS
# ===========================================================================
DEFAULT_BLOCK_START = date(2026, 3, 2)

EXCLUDE_SURNAMES = {
    'Curtin', 'Rouillon', 'Turro', 'Tyler',
    'Wyborn', 'Wagstaffe', 'Pinkerton',
}

# Styling
NAVY = '1F2F4E'
ACCENT = '2E5D9F'
LIGHT = 'EEF2F7'
ALT_ROW = 'F7F9FC'
WHITE = 'FFFFFF'
RED = 'C0392B'
BROWN = '8B4513'
ORANGE = 'D68910'
YELLOW = 'F1C40F'
GREEN = '27AE60'
PURPLE = '8E44AD'

SEG_COLOURS = {
    'S1': RED,
    'S2': BROWN,
    'S3': ORANGE,
    'S4': YELLOW,
    'Active': GREEN,
}

SEG_LABELS = {
    'S1': 'Never Engaged',
    'S2': 'Pre-Block Ghosts',
    'S3': 'W1 Drop-Offs',
    'S4': 'Active Then Absent',
    'Active': 'Currently Active',
}

SEG_DESCRIPTIONS = {
    'S1': 'No record of accessing the subject site during the teaching block.',
    'S2': 'Last access was before the block started — logged in during orientation or pre-teaching but never returned.',
    'S3': 'Last access fell within Week 1. Engaged briefly at the start then disappeared.',
    'S4': 'Was active after Week 1 but has not accessed in the current week.',
    'Active': 'Accessed the subject during the current teaching week.',
}

# ===========================================================================
# CLASS LIST PARSING  (mirrors engagement_report_app.py)
# ===========================================================================
def load_classlist(file_bytes):
    if file_bytes[:2] == b'PK':
        return _load_classlist_xlsx(file_bytes)
    return _load_classlist_xls(file_bytes)


def _load_classlist_xls(file_bytes):
    book = xlrd.open_workbook(file_contents=file_bytes)
    sheet = book.sheet_by_index(0)
    headers = [str(sheet.cell_value(6, c)).strip().lower() for c in range(sheet.ncols)]
    cmap = {h: i for i, h in enumerate(headers) if h}
    if 'student_code' not in cmap:
        raise ValueError("Class list header row not found at row 7. Check the file format.")
    cm = {
        'sid':    cmap['student_code'],
        'first':  cmap.get('first_name', 1),
        'last':   cmap.get('last_name', 2),
        'course': cmap.get('course', 5),
        'email':  cmap.get('email_address', 6),
    }
    subject_raw = str(sheet.cell_value(1, 0)).strip()
    subject_match = re.match(r'([A-Z]+\d+)', subject_raw)
    subject_code = subject_match.group(1) if subject_match else 'UNKNOWN'
    students = {}
    for r in range(7, sheet.nrows):
        sid = str(sheet.cell_value(r, cm['sid'])).strip()
        if not sid or sid == 'student_code':
            continue
        if sid.startswith('30') or sid.startswith('96'):
            continue
        last = str(sheet.cell_value(r, cm['last'])).strip()
        if last in EXCLUDE_SURNAMES:
            continue
        course_val = str(sheet.cell_value(r, cm['course'])).strip()
        students[sid] = {
            'sid': sid,
            'first': str(sheet.cell_value(r, cm['first'])).strip(),
            'last': last,
            'course': course_val,
            'course_code': course_val,
            'discipline_subject': '',
            'discipline_class': '',
            'email': str(sheet.cell_value(r, cm['email'])).strip(),
            'discipline_teacher': '',
        }
    return subject_code, students


def _load_classlist_xlsx(file_bytes):
    wb_cl = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb_cl.active
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            headers[str(val).strip()] = c
    required = {'Student ID', 'First Name', 'Last Name'}
    missing = required - set(headers.keys())
    if missing:
        raise ValueError(f"Enriched class list missing columns: {missing}")
    subject_code = 'UNKNOWN'
    subj_col_name = None
    for candidate in ('GEDU Subject', 'Subject Code'):
        if candidate in headers:
            subj_col_name = candidate
            break
    if subj_col_name:
        raw = str(ws.cell(2, headers[subj_col_name]).value or '').strip()
        m = re.match(r'([A-Z]+\d+)', raw)
        if m:
            subject_code = m.group(1)

    def cell_str(row, col_name, default_col=1):
        c = headers.get(col_name, default_col)
        v = ws.cell(row, c).value
        if v is None:
            return ''
        if isinstance(v, (int, float)):
            return str(int(v))
        return str(v).strip()

    students = {}
    for r in range(2, ws.max_row + 1):
        sid_raw = ws.cell(r, headers['Student ID']).value
        if sid_raw is None:
            continue
        sid = str(int(sid_raw)).strip() if isinstance(sid_raw, (int, float)) else str(sid_raw).strip()
        if not sid or not sid.isdigit():
            continue
        if sid.startswith('30') or sid.startswith('96'):
            continue
        last = cell_str(r, 'Last Name')
        if last in EXCLUDE_SURNAMES:
            continue
        course_code = cell_str(r, 'Course Code')
        students[sid] = {
            'sid': sid,
            'first': cell_str(r, 'First Name'),
            'last': last,
            'course': course_code,
            'course_code': course_code,
            'discipline_subject': cell_str(r, 'Discipline Subject'),
            'discipline_class': cell_str(r, 'Discipline Class'),
            'email': cell_str(r, 'Email'),
            'discipline_teacher': cell_str(r, 'Discipline Teacher'),
        }
    return subject_code, students


# ===========================================================================
# PERFORMANCE DASHBOARD PDF PARSING
# ===========================================================================
def parse_dashboard_pdf(file_bytes):
    """Extract student access data from a Blackboard Performance Dashboard PDF.

    Returns:
        dashboard_date: date the dashboard was exported (or None)
        access_data: dict {username: {'last_access': date|None, 'days_since': int|None, 'never': bool}}
    """
    access_data = {}
    dashboard_date = None

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        full_text = ''
        for page in pdf.pages:
            text = page.extract_text() or ''
            full_text += text + '\n'

    # Try to extract dashboard export date from header
    date_match = re.search(r'(\d{2}/\d{2}/\d{4}),?\s*\d{2}:\d{2}', full_text)
    if date_match:
        dashboard_date = _parse_date(date_match.group(1))

    # Find all Role: positions
    role_pattern = re.compile(
        r'Role:\s*(Student|Instructor|Teaching\s*Assistant|Support|Auditor)',
        re.IGNORECASE,
    )
    positions = []
    for m in role_pattern.finditer(full_text):
        positions.append({'index': m.start(), 'role': re.sub(r'\s+', ' ', m.group(1)).strip()})

    for i, pos in enumerate(positions):
        role = pos['role']
        if role.lower() != 'student':
            continue

        # Chunk: from previous role to next role
        start = positions[i - 1]['index'] if i > 0 else 0
        end = positions[i + 1]['index'] if i < len(positions) - 1 else len(full_text)
        chunk = full_text[start:end]

        # Skip preview users
        if 'previewuser' in chunk.lower() or 'PreviewUser' in chunk:
            continue

        # Extract username
        user_match = re.search(r'Username:\s*(\d{5,10})', chunk)
        if not user_match:
            continue
        username = user_match.group(1)

        # Already seen?
        if username in access_data:
            continue

        # Extract access date
        after_role = chunk[chunk.lower().find('role:'):]
        access_match = re.search(
            r'(?:Last\s*Subject\s*Access:|Access:)\s*(\d{2}/\d{2}/\d{4})\s+\d{1,2}:\d{2}:\d{2}\s*[AP]M',
            after_role, re.IGNORECASE,
        )
        never_match = re.search(
            r'(?:Last\s*Subject\s*Access:|Access:)\s*Never',
            after_role, re.IGNORECASE,
        )

        last_access = None
        days_since = None
        never = False

        if access_match:
            last_access = _parse_date(access_match.group(1))
            if last_access and dashboard_date:
                days_since = (dashboard_date - last_access).days
        elif never_match:
            never = True

        # Try explicit days-since value
        days_match = re.search(
            r'Days\s*Since\s*Last\s*Subject\s*Access\s*[:\s]*(\d+|Never)',
            after_role, re.IGNORECASE,
        )
        if days_match:
            if days_match.group(1).lower() == 'never':
                never = True
            else:
                days_since = int(days_match.group(1))

        access_data[username] = {
            'last_access': last_access,
            'days_since': days_since,
            'never': never,
        }

    return dashboard_date, access_data


def _parse_date(s):
    """Parse DD/MM/YYYY string to date."""
    try:
        parts = s.strip().split('/')
        return date(int(parts[2]), int(parts[1]), int(parts[0]))
    except (ValueError, IndexError):
        return None


# ===========================================================================
# SEGMENTATION
# ===========================================================================
def week_of(dt, block_start):
    """Return the teaching week number (1-based) for a date, or 0 if before block start."""
    diff = (dt - block_start).days
    if diff < 0:
        return 0
    return diff // 7 + 1


def week_range(week_num, block_start):
    """Return (start_date, end_date) for a given teaching week."""
    start = block_start + timedelta(days=(week_num - 1) * 7)
    end = start + timedelta(days=6)
    return start, end


def segment_students(students, access_data, block_start, dashboard_date):
    """Assign each student a segment based on last access date.

    Returns:
        results: list of dicts with student info + segment data
        current_week: int
        days_into_week: int
        counts: dict of segment counts
    """
    current_week = week_of(dashboard_date, block_start)
    if current_week < 1:
        current_week = 1
    days_into_week = (dashboard_date - block_start).days % 7 + 1
    is_partial = days_into_week < 7

    results = []
    counts = {'S1': 0, 'S2': 0, 'S3': 0, 'S4_JD': 0, 'S4_LS': 0, 'Active': 0}

    for sid, st in students.items():
        acc = access_data.get(sid)

        row = {
            'sid': sid,
            'first': st['first'],
            'last': st['last'],
            'course': st.get('course_code') or st.get('course', ''),
            'discipline_subject': st.get('discipline_subject', ''),
            'discipline_class': st.get('discipline_class', ''),
            'discipline_teacher': st.get('discipline_teacher', ''),
            'email': st.get('email', ''),
            'last_access': None,
            'days_since': None,
            'last_week': None,
            'segment': None,
            'sub_group': '',
            'in_dashboard': acc is not None,
        }

        # Not in dashboard at all, or explicitly Never
        if acc is None or acc['never']:
            row['segment'] = 'S1'
            row['last_access'] = None
            row['days_since'] = None
            counts['S1'] += 1
            results.append(row)
            continue

        row['last_access'] = acc['last_access']
        row['days_since'] = acc['days_since']

        if acc['last_access']:
            wk = week_of(acc['last_access'], block_start)
            row['last_week'] = wk

            if wk == 0:
                # Before block start
                row['segment'] = 'S2'
                counts['S2'] += 1
            elif wk == 1 and current_week > 1:
                # W1 drop-off
                row['segment'] = 'S3'
                counts['S3'] += 1
            elif wk < current_week and wk > 1:
                # Active then absent
                row['segment'] = 'S4'
                if wk == current_week - 1:
                    row['sub_group'] = 'Just Dropped'
                    counts['S4_JD'] += 1
                else:
                    row['sub_group'] = 'Long Silent'
                    counts['S4_LS'] += 1
            else:
                # Current week or later
                row['segment'] = 'Active'
                counts['Active'] += 1
        else:
            # Has record but no parseable date — treat as S1
            row['segment'] = 'S1'
            counts['S1'] += 1

        results.append(row)

    return results, current_week, days_into_week, is_partial, counts


# ===========================================================================
# WORKBOOK STYLING HELPERS  (mirrors engagement_report_app.py)
# ===========================================================================
def thin_border():
    side = Side(style='thin', color='D5DBDB')
    return Border(left=side, right=side, top=side, bottom=side)


def write_tab_header(ws, title, subtitle, description, n_cols, colour=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1, title)
    c.font = Font(name='Arial', size=14, bold=True, color=WHITE)
    c.fill = PatternFill('solid', start_color=NAVY)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 26

    fill_colour = colour or ACCENT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws.cell(2, 1, subtitle or '')
    c.font = Font(name='Arial', size=11, bold=True, color=WHITE)
    c.fill = PatternFill('solid', start_color=fill_colour)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 20

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    c = ws.cell(3, 1, description)
    c.font = Font(name='Arial', size=10, italic=True, color='2C3E50')
    c.fill = PatternFill('solid', start_color=LIGHT)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    ws.row_dimensions[3].height = 36
    ws.row_dimensions[4].height = 6


def write_col_headers(ws, headers, row=5):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.font = Font(name='Arial', size=10, bold=True, color=WHITE)
        c.fill = PatternFill('solid', start_color=ACCENT)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 30


def write_data_rows(ws, data_rows, start_row=6, seg_col=None):
    for ri, row in enumerate(data_rows):
        excel_row = start_row + ri
        fill = PatternFill('solid', start_color=ALT_ROW) if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(excel_row, ci, val)
            c.font = Font(name='Arial', size=10)
            if fill:
                c.fill = fill
            c.alignment = Alignment(
                horizontal='left' if ci <= 5 else 'center',
                vertical='center',
            )
            c.border = thin_border()


def write_seg_badge(ws, row, col, segment, sub_group=''):
    c = ws.cell(row, col)
    colour = SEG_COLOURS.get(segment, ACCENT)
    if segment == 'S4' and sub_group == 'Long Silent':
        colour = PURPLE
    c.fill = PatternFill('solid', start_color=colour)
    c.font = Font(name='Arial', size=10, bold=True, color=WHITE)


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def fmt_date(dt):
    if dt is None:
        return 'Never'
    return dt.strftime('%d/%m/%Y')


def fmt_week(wk):
    if wk is None:
        return '—'
    if wk == 0:
        return 'Pre-block'
    return f'W{wk}'


# ===========================================================================
# WORKBOOK BUILDER
# ===========================================================================
def build_workbook(subject_code, results, current_week, days_into_week, is_partial,
                   block_start, dashboard_date, counts, enrolled):
    wb = Workbook()
    wb.remove(wb.active)

    s4_total = counts['S4_JD'] + counts['S4_LS']
    active_count = counts['Active']
    at_risk = enrolled - active_count

    partial_tag = f' ({days_into_week}d partial)' if is_partial else ''
    ws_start, ws_end = week_range(current_week, block_start)

    # ── Summary ──
    ws = wb.create_sheet('Summary')
    write_tab_header(
        ws,
        f'{subject_code} — Engagement Pulse Check — W{current_week}{partial_tag}',
        f'Block start: {block_start.strftime("%b %-d %Y")}  •  Dashboard: {dashboard_date.strftime("%b %-d %Y")}  •  Enrolled: {enrolled}',
        ('Approximate segmentation based on last-access date only. '
         'S5–S7 (late arrivals, fading, sustained) require weekly hit-count data from usage reports — '
         'use the full Engagement Report Builder for those.'),
        7,
    )

    seg_meta = [
        ('S1', 'Never Engaged', counts['S1'], SEG_DESCRIPTIONS['S1']),
        ('S2', 'Pre-Block Ghosts', counts['S2'], SEG_DESCRIPTIONS['S2']),
        ('S3', 'W1 Drop-Offs', counts['S3'], SEG_DESCRIPTIONS['S3']),
        ('S4', 'Active Then Absent', s4_total, SEG_DESCRIPTIONS['S4']),
        ('', '  ↳ Just Dropped', counts['S4_JD'], f'Last access in W{current_week - 1}.' if current_week > 1 else ''),
        ('', '  ↳ Long Silent', counts['S4_LS'], 'Last access 2+ weeks ago.'),
        ('Active', 'Currently Active', active_count, SEG_DESCRIPTIONS['Active']),
    ]

    write_col_headers(ws, ['Segment', 'Label', 'Count', '% of Enrolled', 'Status', 'Description', ''], row=5)
    for ri, (code, label, count, desc) in enumerate(seg_meta):
        excel_row = 6 + ri
        pct = f'{count / enrolled * 100:.1f}%' if enrolled > 0 else '0%'
        status = ''
        if code in ('S1', 'S2'):
            status = 'Critical'
        elif code == 'S3':
            status = 'High Risk'
        elif code == 'S4':
            status = 'Watch'
        elif code == 'Active':
            status = 'Healthy'
        row_data = [code, label, count, pct, status, desc, '']
        fill = PatternFill('solid', start_color=ALT_ROW) if ri % 2 == 0 else None
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(excel_row, ci, val)
            c.font = Font(name='Arial', size=10, bold=(code != ''))
            if fill:
                c.fill = fill
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(ci == 6))
            c.border = thin_border()
        if code and code != 'Active':
            ws.cell(excel_row, 1).fill = PatternFill('solid', start_color=SEG_COLOURS.get(code, ACCENT))
            ws.cell(excel_row, 1).font = Font(name='Arial', size=10, bold=True, color=WHITE)
        elif code == 'Active':
            ws.cell(excel_row, 1).fill = PatternFill('solid', start_color=GREEN)
            ws.cell(excel_row, 1).font = Font(name='Arial', size=10, bold=True, color=WHITE)

    # Total row
    total_row = 6 + len(seg_meta)
    for ci, val in enumerate(['TOTAL', '', enrolled, '100%', '', '', ''], 1):
        c = ws.cell(total_row, ci, val)
        c.font = Font(name='Arial', size=10, bold=True)
        c.fill = PatternFill('solid', start_color=LIGHT)
        c.border = thin_border()
        c.alignment = Alignment(horizontal='left', vertical='center')

    # Notes
    note_row = total_row + 2
    notes = [
        f'Teaching week: W{current_week} ({ws_start.strftime("%b %-d")} – {ws_end.strftime("%-d")}){"  ⚠ PARTIAL — " + str(days_into_week) + " days of data" if is_partial else ""}',
        f'At risk (S1+S2+S3+S4): {at_risk} ({at_risk / enrolled * 100:.1f}%)',
        f'Students in class list but not in dashboard: {sum(1 for r in results if not r["in_dashboard"])}',
        'This report uses last-access date only. For full S5–S7 segmentation (late arrivals, fading, sustained), '
        'use the Engagement Report Builder with usage + login report files.',
    ]
    if is_partial:
        notes.insert(1,
            f'PARTIAL WEEK: W{current_week} has {days_into_week} days. S4 may be inflated '
            f'(students who have not yet accessed this week will appear as absent).')
    for i, note in enumerate(notes):
        c = ws.cell(note_row + i, 1, note)
        ws.merge_cells(start_row=note_row + i, start_column=1, end_row=note_row + i, end_column=7)
        is_warning = ('PARTIAL' in note or '⚠' in note)
        c.font = Font(name='Arial', size=9, italic=is_warning, color=(RED if is_warning else '2C3E50'), bold=is_warning)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[note_row + i].height = 32 if is_warning else 18

    autosize(ws, [10, 22, 10, 14, 12, 60, 2])
    ws.freeze_panes = 'A6'

    # ── Segment detail sheets ──
    col_headers = [
        'Surname', 'First Name', 'Student ID', 'Course',
        'Disc. Class', 'Disc. Teacher', 'Email',
        'Last Access', 'Days Since', 'Last Active Wk', 'Segment',
    ]
    col_widths = [22, 18, 14, 12, 12, 20, 38, 14, 12, 14, 18]

    def _make_sheet(sheet_name, title, subtitle, desc, segment_filter, colour,
                    sub_filter=None, sort_key=None):
        filtered = [r for r in results if r['segment'] == segment_filter]
        if sub_filter:
            filtered = [r for r in filtered if r['sub_group'] == sub_filter]
        if sort_key:
            filtered.sort(key=sort_key)
        else:
            filtered.sort(key=lambda r: (r['last'] or '').lower())

        ws_s = wb.create_sheet(sheet_name[:31])
        write_tab_header(ws_s, title, f'{len(filtered)} students', desc, len(col_headers), colour)
        write_col_headers(ws_s, col_headers, row=5)

        for ri, r in enumerate(filtered):
            excel_row = 6 + ri
            seg_label = r['sub_group'] if r['sub_group'] else SEG_LABELS.get(r['segment'], r['segment'])
            row_data = [
                r['last'], r['first'], r['sid'], r['course'],
                r['discipline_class'], r['discipline_teacher'], r['email'],
                fmt_date(r['last_access']), r['days_since'] if r['days_since'] is not None else '—',
                fmt_week(r['last_week']), seg_label,
            ]
            fill = PatternFill('solid', start_color=ALT_ROW) if ri % 2 == 0 else None
            for ci, val in enumerate(row_data, 1):
                c = ws_s.cell(excel_row, ci, val)
                c.font = Font(name='Arial', size=10)
                if fill:
                    c.fill = fill
                c.alignment = Alignment(
                    horizontal='left' if ci <= 7 else 'center',
                    vertical='center',
                )
                c.border = thin_border()
            write_seg_badge(ws_s, excel_row, len(col_headers), r['segment'], r['sub_group'])

        autosize(ws_s, col_widths)
        ws_s.freeze_panes = 'A6'
        return ws_s

    # S1
    _make_sheet('S1 Never Engaged', 'S1 — Never Engaged', '', SEG_DESCRIPTIONS['S1'], 'S1', RED)

    # S2
    _make_sheet('S2 Pre-Block Ghosts', 'S2 — Pre-Block Ghosts', '', SEG_DESCRIPTIONS['S2'], 'S2', BROWN,
                sort_key=lambda r: (r['days_since'] or 9999) * -1)

    # S3
    _make_sheet('S3 W1 Drop-Offs', 'S3 — W1 Drop-Offs', '', SEG_DESCRIPTIONS['S3'], 'S3', ORANGE,
                sort_key=lambda r: (r['days_since'] or 9999) * -1)

    # S4 — combined with sub-group headers
    s4_all = sorted(
        [r for r in results if r['segment'] == 'S4'],
        key=lambda r: (0 if r['sub_group'] == 'Just Dropped' else 1, -(r['days_since'] or 0)),
    )
    ws_s4 = wb.create_sheet('S4 Active Then Absent'[:31])
    write_tab_header(
        ws_s4,
        f'S4 — Active Then Absent in W{current_week}',
        f'{len(s4_all)} total  •  Just Dropped: {counts["S4_JD"]}  •  Long Silent: {counts["S4_LS"]}',
        SEG_DESCRIPTIONS['S4'],
        len(col_headers),
        YELLOW,
    )
    write_col_headers(ws_s4, col_headers, row=5)

    current_row = 6
    jd_list = [r for r in s4_all if r['sub_group'] == 'Just Dropped']
    ls_list = [r for r in s4_all if r['sub_group'] == 'Long Silent']

    def _write_subgroup(ws_s4, group_list, group_label, banner_colour, start_row):
        if not group_list:
            return start_row
        # Banner row
        ws_s4.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(col_headers))
        c = ws_s4.cell(start_row, 1, f'▼ {group_label} ({len(group_list)} students)')
        c.font = Font(name='Arial', size=10, bold=True, color=WHITE)
        c.fill = PatternFill('solid', start_color=banner_colour)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws_s4.row_dimensions[start_row].height = 22
        start_row += 1

        for ri, r in enumerate(group_list):
            seg_label = r['sub_group']
            row_data = [
                r['last'], r['first'], r['sid'], r['course'],
                r['discipline_class'], r['discipline_teacher'], r['email'],
                fmt_date(r['last_access']), r['days_since'] if r['days_since'] is not None else '—',
                fmt_week(r['last_week']), seg_label,
            ]
            fill = PatternFill('solid', start_color=ALT_ROW) if ri % 2 == 0 else None
            for ci, val in enumerate(row_data, 1):
                c = ws_s4.cell(start_row, ci, val)
                c.font = Font(name='Arial', size=10)
                if fill:
                    c.fill = fill
                c.alignment = Alignment(
                    horizontal='left' if ci <= 7 else 'center',
                    vertical='center',
                )
                c.border = thin_border()
            write_seg_badge(ws_s4, start_row, len(col_headers), 'S4', r['sub_group'])
            start_row += 1
        return start_row

    current_row = _write_subgroup(ws_s4, jd_list,
        f'JUST DROPPED — Last access in W{current_week - 1}', RED, current_row)
    current_row = _write_subgroup(ws_s4, ls_list,
        f'LONG SILENT — Last access W2–W{max(current_week - 2, 2)}', PURPLE, current_row)
    autosize(ws_s4, col_widths)
    ws_s4.freeze_panes = 'A6'

    # Active
    _make_sheet('Active', 'Currently Active', '', SEG_DESCRIPTIONS['Active'], 'Active', GREEN)

    # ── All Students ──
    all_sorted = sorted(results, key=lambda r: (
        {'S1': 0, 'S2': 1, 'S3': 2, 'S4': 3, 'Active': 4}.get(r['segment'], 9),
        (r['last'] or '').lower(),
    ))
    ws_all = wb.create_sheet('All Students')
    write_tab_header(
        ws_all,
        f'{subject_code} — All Students',
        f'{enrolled} enrolled  •  W{current_week}{partial_tag}',
        'All students sorted by segment then surname.',
        len(col_headers),
    )
    write_col_headers(ws_all, col_headers, row=5)
    for ri, r in enumerate(all_sorted):
        excel_row = 6 + ri
        seg_label = r['sub_group'] if r['sub_group'] else SEG_LABELS.get(r['segment'], r['segment'])
        row_data = [
            r['last'], r['first'], r['sid'], r['course'],
            r['discipline_class'], r['discipline_teacher'], r['email'],
            fmt_date(r['last_access']), r['days_since'] if r['days_since'] is not None else '—',
            fmt_week(r['last_week']), seg_label,
        ]
        fill = PatternFill('solid', start_color=ALT_ROW) if ri % 2 == 0 else None
        for ci, val in enumerate(row_data, 1):
            c = ws_all.cell(excel_row, ci, val)
            c.font = Font(name='Arial', size=10)
            if fill:
                c.fill = fill
            c.alignment = Alignment(
                horizontal='left' if ci <= 7 else 'center',
                vertical='center',
            )
            c.border = thin_border()
        write_seg_badge(ws_all, excel_row, len(col_headers), r['segment'], r['sub_group'])

    autosize(ws_all, col_widths)
    ws_all.freeze_panes = 'A6'

    return wb


# ===========================================================================
# STREAMLIT UI
# ===========================================================================
st.set_page_config(page_title='Engagement Pulse Check', layout='wide', page_icon='📡')
st.title('📡 Engagement Pulse Check')
st.caption(
    'Quick engagement segmentation from a class list + Performance Dashboard PDF. '
    'No usage or login report files required.'
)

with st.expander('How it works', expanded=False):
    st.markdown("""
**Inputs:**
- **Class list** (.xls from Blackboard or enriched .xlsx) — source of truth for enrolled students.
- **Performance Dashboard PDF** — exported from Blackboard with **Show All** enabled.

**Segments** (based on last-access date relative to block start):

| Segment | Rule |
|---|---|
| **S1 — Never Engaged** | Never accessed the subject |
| **S2 — Pre-Block Ghosts** | Last access before block started |
| **S3 — W1 Drop-Offs** | Last access in Week 1 only |
| **S4 — Active Then Absent** | Last access after W1 but not in current week |
| ↳ Just Dropped | Last access was previous week |
| ↳ Long Silent | Last access 2+ weeks ago |
| **Active** | Accessed during current week |

**Limitations:** This tool cannot distinguish S5 (late arrivals), S6 (fading), or S7 (sustained) —
those require weekly hit-count data from usage reports. Use the full **Engagement Report Builder**
for that level of detail.
""")

# ── Sidebar controls ──
with st.sidebar:
    st.header('Settings')
    block_start = st.date_input(
        'Block start date',
        value=DEFAULT_BLOCK_START,
        help='First day of Week 1 teaching. All week numbering is anchored to this date.',
    )

# ── File uploaders ──
col1, col2 = st.columns(2)
with col1:
    classlist_file = st.file_uploader(
        'Class list (.xls / .xlsx)',
        type=['xls', 'xlsx'],
        key='cl',
        help='Blackboard class list or enriched version with Course Code, Discipline Class, etc.',
    )
with col2:
    dashboard_file = st.file_uploader(
        'Performance Dashboard (.pdf)',
        type=['pdf'],
        key='pd',
        help='Export from Blackboard with "Show All" enabled, saved as PDF.',
    )

run_btn = st.button(
    'Generate pulse check',
    type='primary',
    disabled=not (classlist_file and dashboard_file),
)

if run_btn:
    try:
        # ── Load class list ──
        with st.spinner('Loading class list…'):
            subject_code, students = load_classlist(classlist_file.getvalue())
        st.success(f'**{subject_code}** — {len(students)} enrolled (after exclusions)')

        # ── Parse dashboard PDF ──
        with st.spinner('Parsing Performance Dashboard…'):
            dashboard_date, access_data = parse_dashboard_pdf(dashboard_file.getvalue())

        if dashboard_date is None:
            dashboard_date = date.today()
            st.warning('Could not detect dashboard export date — using today.')
        else:
            st.info(f'Dashboard exported: **{dashboard_date.strftime("%b %-d %Y")}**')

        dashboard_students = sum(1 for k in access_data if k in students)
        st.write(
            f'Dashboard: **{len(access_data)}** student records parsed, '
            f'**{dashboard_students}** matched to class list.'
        )

        # ── Segment ──
        with st.spinner('Segmenting students…'):
            results, current_week, days_into_week, is_partial, counts = segment_students(
                students, access_data, block_start, dashboard_date,
            )

        s4_total = counts['S4_JD'] + counts['S4_LS']
        enrolled = len(students)
        active_count = counts['Active']

        # ── Week info ──
        ws_start, ws_end = week_range(current_week, block_start)
        partial_msg = f' ({days_into_week}d partial)' if is_partial else ''
        st.success(
            f'Current week: **W{current_week}**{partial_msg} '
            f'({ws_start.strftime("%b %-d")} – {ws_end.strftime("%-d")})'
        )
        if is_partial:
            st.warning(
                f'W{current_week} has {days_into_week} days of data. '
                f'S4 may be inflated — students who haven\'t accessed yet this week '
                f'will show as absent.'
            )

        # ── Metrics ──
        st.subheader('Segment overview')
        cols = st.columns(5)
        seg_display = [
            ('S1', 'Never', counts['S1'], RED),
            ('S2', 'Ghosts', counts['S2'], BROWN),
            ('S3', 'W1 Drop', counts['S3'], ORANGE),
            ('S4', 'Absent', s4_total, YELLOW),
            ('Active', 'Active', active_count, GREEN),
        ]
        for i, (code, label, count, colour) in enumerate(seg_display):
            pct = f'{count / enrolled * 100:.1f}%' if enrolled > 0 else '0%'
            cols[i].metric(f'{code} {label}', count, pct)

        # ── At-risk bar ──
        at_risk = enrolled - active_count
        st.progress(active_count / enrolled if enrolled > 0 else 0)
        st.caption(
            f'**{active_count}** active ({active_count / enrolled * 100:.1f}%) · '
            f'**{at_risk}** at risk ({at_risk / enrolled * 100:.1f}%)'
        )

        # ── Detail tables ──
        st.subheader('Flagged students')

        def show_table(label, segment, sub_group=None, colour=RED):
            filtered = [r for r in results if r['segment'] == segment]
            if sub_group:
                filtered = [r for r in filtered if r['sub_group'] == sub_group]
            if not filtered:
                return
            filtered.sort(key=lambda r: -(r['days_since'] or 9999))
            with st.expander(f'{label} ({len(filtered)})', expanded=(len(filtered) > 0 and len(filtered) <= 20)):
                table_data = []
                for r in filtered:
                    table_data.append({
                        'Name': f'{r["last"]}, {r["first"]}',
                        'SID': r['sid'],
                        'Course': r['course'],
                        'Class': r['discipline_class'],
                        'Last Access': fmt_date(r['last_access']),
                        'Days Since': r['days_since'] if r['days_since'] is not None else '—',
                        'Last Wk': fmt_week(r['last_week']),
                    })
                st.dataframe(table_data, use_container_width=True, hide_index=True)

        show_table('S1 — Never Engaged', 'S1')
        show_table('S2 — Pre-Block Ghosts', 'S2')
        show_table('S3 — W1 Drop-Offs', 'S3')
        show_table(f'S4 — Just Dropped (W{current_week - 1} → absent)', 'S4', 'Just Dropped')
        show_table('S4 — Long Silent', 'S4', 'Long Silent')

        not_in_dashboard = sum(1 for r in results if not r['in_dashboard'])
        if not_in_dashboard > 0:
            st.info(
                f'**{not_in_dashboard}** students in the class list were not found in the '
                f'Performance Dashboard. These are counted as S1 (Never Engaged). '
                f'If the dashboard was not exported with "Show All", some may be missing.'
            )

        # ── Build workbook ──
        with st.spinner('Building report…'):
            wb = build_workbook(
                subject_code, results, current_week, days_into_week,
                is_partial, block_start, dashboard_date, counts, enrolled,
            )
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

        date_str = dashboard_date.strftime('%Y%m%d')
        suffix = '_Partial' if is_partial else ''
        filename = f'{subject_code}_Pulse_Check_W{current_week}_{date_str}{suffix}.xlsx'

        st.download_button(
            '⬇ Download report',
            data=buf,
            file_name=filename,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            type='primary',
        )

    except Exception as e:
        st.error(f'Error: {e}')
        import traceback
        st.code(traceback.format_exc())
