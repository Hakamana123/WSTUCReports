import io
import re
import zipfile
from collections import defaultdict

import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink

st.set_page_config(page_title="File Splitter", layout="wide")

st.title("File Splitter")
st.caption(
    "Upload a finished Class Report (from the Engagement Report builder) and split it into "
    "one workbook per program — each containing only that program's class sheets, ready to "
    "send straight to the program coordinator."
)

HYPERLINK_SHEET_RE = re.compile(r"^#'?(.+?)'?!")
INDEX_COLUMNS = ["#", "Class", "Teacher", "Enrolled", "Active", "At Risk", "AS1 Sub", "AS2 Sub", "AS3 Sub"]


def _resolve_hyperlink_sheet(cell):
    if not cell.hyperlink or not cell.hyperlink.target:
        return None
    m = HYPERLINK_SHEET_RE.match(cell.hyperlink.target)
    return m.group(1) if m else None


def parse_class_index(wb):
    """Read the Class Index sheet and return one entry per class row.

    Returns None if the workbook has no Class Index sheet with a Program column
    (i.e. not a class report produced by the Engagement Report builder).
    """
    if "Class Index" not in wb.sheetnames:
        return None
    ws = wb["Class Index"]

    header_row = None
    for r in range(1, min(15, ws.max_row) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if vals and vals[0] == "#" and "Program" in vals:
            header_row = r
            break
    if header_row is None:
        return None

    entries = []
    for r in range(header_row + 1, ws.max_row + 1):
        num_val = ws.cell(row=r, column=1).value
        if not isinstance(num_val, (int, float)):
            continue  # group header / blank row, not a class row
        class_cell = ws.cell(row=r, column=2)
        if class_cell.value == "No match":
            continue  # unmatched students, not a real class — never split into a program file
        sheet_name = _resolve_hyperlink_sheet(class_cell) or str(class_cell.value)
        entries.append(
            {
                "sheet": sheet_name,
                "class_label": class_cell.value,
                "program": "" if ws.cell(row=r, column=3).value is None else str(ws.cell(row=r, column=3).value),
                "teacher": ws.cell(row=r, column=4).value,
                "enrolled": ws.cell(row=r, column=5).value,
                "active": ws.cell(row=r, column=6).value,
                "at_risk": ws.cell(row=r, column=7).value,
                "as1_sub": ws.cell(row=r, column=8).value,
                "as2_sub": ws.cell(row=r, column=9).value,
                "as3_sub": ws.cell(row=r, column=10).value,
            }
        )
    return entries


def _write_program_index(wb, program, prog_entries):
    """Build a fresh 'Class Index' sheet listing just this program's classes.

    Built from scratch rather than derived by deleting rows from the original:
    openpyxl's delete_rows does not reliably carry hyperlinks along with the
    rows it shifts, which corrupts a sheet with as many rows as Class Index
    has. A freshly-written sheet has no such risk. It keeps the name
    "Class Index" so each retained class sheet's "Back to Index" link
    (which points at '#'Class Index'!A1') keeps working.
    """
    ws = wb.create_sheet("Class Index", 0)

    ws.cell(row=1, column=1, value=f"Program {program} — Class Index").font = Font(bold=True, size=14)
    ws.cell(
        row=2,
        column=1,
        value=f"{len(prog_entries)} class(es)  •  Click a class name to jump to its sheet",
    ).font = Font(italic=True, color="666666")

    header_row = 4
    header_fill = PatternFill("solid", start_color="1F2937", end_color="1F2937")
    for col, title in enumerate(INDEX_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for i, e in enumerate(sorted(prog_entries, key=lambda e: e["class_label"] or ""), start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=i)
        class_cell = ws.cell(row=r, column=2, value=e["class_label"])
        if e["sheet"] in wb.sheetnames:
            class_cell.hyperlink = Hyperlink(ref=class_cell.coordinate, location=f"'{e['sheet']}'!A1")
            class_cell.font = Font(color="0563C1", underline="single")
        ws.cell(row=r, column=3, value=e["teacher"])
        ws.cell(row=r, column=4, value=e["enrolled"])
        ws.cell(row=r, column=5, value=e["active"])
        ws.cell(row=r, column=6, value=e["at_risk"])
        ws.cell(row=r, column=7, value=e["as1_sub"])
        ws.cell(row=r, column=8, value=e["as2_sub"])
        ws.cell(row=r, column=9, value=e["as3_sub"])

    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "A5"


def build_program_workbook(file_bytes, program, prog_entries, all_class_sheets):
    """Load a fresh copy of the report and strip it down to just this program.

    Only whole-sheet deletion is used on the original workbook (safe, well
    supported by openpyxl) — kept class sheets are byte-for-byte untouched, so
    all their formatting, conditional formatting and hyperlinks survive. The
    old Class Index, Summary and Assessment Detail sheets mix in every other
    program's data and aren't safely row-filterable (see _write_program_index
    docstring), so they're dropped and replaced with a fresh, program-scoped
    Class Index built from parsed data rather than sheet surgery.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    keep_sheets = {e["sheet"] for e in prog_entries}

    for name in list(wb.sheetnames):
        if name in ("No match", "Summary", "Assessment Detail", "Class Index"):
            del wb[name]
        elif name in all_class_sheets and name not in keep_sheets:
            del wb[name]

    _write_program_index(wb, program, prog_entries)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


uploaded = st.file_uploader("Class Report (.xlsx)", type=["xlsx"])

if uploaded:
    file_bytes = uploaded.getvalue()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    entries = parse_class_index(wb)

    if entries is None:
        st.error(
            "Couldn't find a 'Class Index' sheet with a Program column in this file. "
            "Upload a Class Report generated by the Engagement Report builder."
        )
    elif not entries:
        st.warning("Class Index has no class rows to split.")
    else:
        by_program = defaultdict(list)
        for e in entries:
            by_program[e["program"] or "(no program)"].append(e)
        all_class_sheets = {e["sheet"] for e in entries}

        st.subheader("Programs found")
        summary_rows = [
            {
                "Program": program,
                "Classes": len(prog_entries),
                "Teachers": len({e["teacher"] for e in prog_entries if e["teacher"]}),
            }
            for program, prog_entries in sorted(by_program.items())
        ]
        st.dataframe(summary_rows, use_container_width=True, hide_index=True)

        base_name = re.sub(r"\.xlsx$", "", uploaded.name, flags=re.IGNORECASE)

        if st.button(f"Build {len(by_program)} program files", type="primary"):
            outputs = {}
            progress = st.progress(0.0)
            for i, (program, prog_entries) in enumerate(sorted(by_program.items())):
                safe_program = re.sub(r"[^A-Za-z0-9_-]+", "_", program)
                buf = build_program_workbook(file_bytes, program, prog_entries, all_class_sheets)
                outputs[f"{base_name}_Program_{safe_program}.xlsx"] = buf
                progress.progress((i + 1) / len(by_program))
            progress.empty()

            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for filename, buf in outputs.items():
                    zf.writestr(filename, buf.getvalue())
            zip_buf.seek(0)

            st.success(f"Built {len(outputs)} program files.")
            st.download_button(
                "Download all as .zip",
                data=zip_buf,
                file_name=f"{base_name}_by_program.zip",
                mime="application/zip",
            )

            with st.expander("Download individual files"):
                for filename, buf in outputs.items():
                    st.download_button(
                        filename,
                        data=buf,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{filename}",
                    )
else:
    st.info("Upload a Class Report to split it by program.")
