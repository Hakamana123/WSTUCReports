"""
Engagement Trend
================
Companion page for the WSUTC Engagement Report Builder.

Reads the "Summary" tab from engagement report workbooks you've already
generated (aggregate segment counts only - no student names, IDs, or emails)
and plots segmentation over time.

Nothing is stored on the server. History lives in a small CSV file that YOU
keep: download it here, and re-upload it next time to keep building the
trend. You can also drop in several past report files at once to backfill.

To add this as a page: place this file in your app's `pages/` folder
(Streamlit auto-discovers it; rename with a number prefix to control its
position in the sidebar, e.g. `2_Engagement_Trend.py`).

Requires: streamlit, pandas, openpyxl (altair ships bundled with streamlit - no extra install needed)
"""
import io
import re
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import altair as alt

st.set_page_config(page_title='Engagement Trend', page_icon='📈', layout='wide')
st.title('Engagement Trend')
st.caption(
    "Nothing here is stored on the server - all processing happens for this "
    "visit only. Upload the Summary tab from your engagement report(s) below, "
    "optionally bring in a trend history CSV from a previous visit, then "
    "download the updated history at the bottom to keep building the picture "
    "next time."
)

SEG_ORDER = ['S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7']
HIST_COLUMNS = ['subject_code', 'snapshot_date', 'snapshot_label'] + SEG_ORDER + ['enrolled']


def parse_report_summary(file_bytes):
    """Extract subject code, snapshot date/label, and S1-S7 counts from the
    Summary tab of an engagement report workbook. Raises ValueError with a
    human-readable message if the file doesn't look like an engagement report."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if 'Summary' not in wb.sheetnames:
        raise ValueError("no 'Summary' tab found - is this an engagement report workbook?")
    ws = wb['Summary']

    title = ws.cell(1, 1).value or ''
    subtitle = ws.cell(2, 1).value or ''

    m_subject = re.match(r'(\S+)\s+Engagement Report\s+—\s+(.+)', title)
    if not m_subject:
        raise ValueError(f"couldn't parse the title row ({title!r}) - is this the main engagement report (not the program/class report)?")
    subject_code, snapshot_label = m_subject.group(1), m_subject.group(2)

    m_date = re.search(r'Latest data:\s*([A-Za-z]+ \d{1,2} \d{4})', subtitle)
    if not m_date:
        raise ValueError(f"couldn't find a 'Latest data' date in the subtitle row ({subtitle!r})")
    snapshot_date = datetime.strptime(m_date.group(1), '%b %d %Y').date()

    m_enrolled = re.search(r'Enrolled\s+(\d+)', subtitle)
    enrolled = int(m_enrolled.group(1)) if m_enrolled else None

    counts = {}
    for r in range(1, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if code in SEG_ORDER:
            val = ws.cell(r, 3).value
            counts[code] = int(val) if isinstance(val, (int, float)) else 0
    if len(counts) != 7:
        raise ValueError(f'found {len(counts)}/7 segment rows in the Summary tab - the sheet layout may have changed')

    if enrolled is None:
        enrolled = sum(counts.values())

    return {
        'subject_code': subject_code,
        'snapshot_date': snapshot_date,
        'snapshot_label': snapshot_label,
        **counts,
        'enrolled': enrolled,
    }


st.subheader('1. Add snapshot(s)')
report_files = st.file_uploader(
    'Upload one or more engagement report workbooks (.xlsx) - only the Summary tab is read. '
    'You can drop in several past reports at once to backfill history.',
    type=['xlsx'], accept_multiple_files=True,
)

st.subheader('2. Bring in your trend history (optional)')
history_file = st.file_uploader(
    'Upload a trend history CSV you downloaded from a previous visit, if you have one',
    type=['csv'],
)

history_df = pd.DataFrame(columns=HIST_COLUMNS)
if history_file is not None:
    try:
        history_df = pd.read_csv(io.StringIO(history_file.getvalue().decode('utf-8')))
        missing_cols = set(HIST_COLUMNS) - set(history_df.columns)
        if missing_cols:
            st.error(f'History file is missing expected columns: {missing_cols}')
            history_df = pd.DataFrame(columns=HIST_COLUMNS)
    except Exception as e:
        st.error(f'Could not read history file: {e}')

new_rows = []
if report_files:
    for f in report_files:
        try:
            new_rows.append(parse_report_summary(f.getvalue()))
        except Exception as e:
            st.error(f'{f.name}: {e}')

if not history_df.empty:
    history_df['snapshot_date'] = pd.to_datetime(history_df['snapshot_date'])

if new_rows:
    new_df = pd.DataFrame(new_rows)
    new_df['snapshot_date'] = pd.to_datetime(new_df['snapshot_date'])
    combined = pd.concat([history_df, new_df], ignore_index=True)
else:
    combined = history_df

if combined.empty:
    st.info('Upload at least one engagement report to get started.')
    st.stop()

# Same subject + same snapshot date re-uploaded -> keep the most recent upload, not a duplicate point
combined = combined.drop_duplicates(subset=['subject_code', 'snapshot_date'], keep='last')
combined = combined.sort_values(['subject_code', 'snapshot_date']).reset_index(drop=True)

st.subheader('3. Trend')
subjects = sorted(combined['subject_code'].unique())
subject = st.selectbox('Subject', subjects)
sub_df = combined[combined['subject_code'] == subject].sort_values('snapshot_date').copy()

sub_df['red'] = sub_df['S1'] + sub_df['S2'] + sub_df['S3']
sub_df['green'] = sub_df['S4'] + sub_df['S6']
sub_df['blue'] = sub_df['S5'] + sub_df['S7']
for c in ['red', 'green', 'blue']:
    sub_df[c + '_pct'] = 100 * sub_df[c] / sub_df['enrolled']

bucket_labels = {
    'red_pct': 'S1+S2+S3 (never engaged / ghosts / drop-offs)',
    'green_pct': 'S4+S6 (active then absent / fading)',
    'blue_pct': 'S5+S7 (late arrivals / sustained)',
}
snapshot_order = list(sub_df['snapshot_label'])

plot_df = sub_df.melt(
    id_vars=['snapshot_label'],
    value_vars=list(bucket_labels.keys()),
    var_name='bucket', value_name='pct',
)
plot_df['bucket'] = plot_df['bucket'].map(bucket_labels)
order_map = {v: i for i, v in enumerate(bucket_labels.values())}
plot_df['bucket_order'] = plot_df['bucket'].map(order_map)

color_scale = alt.Scale(domain=list(bucket_labels.values()), range=['#e34948', '#008300', '#2a78d6'])

bar = alt.Chart(plot_df).mark_bar().encode(
    x=alt.X('snapshot_label:N', sort=snapshot_order, title=None),
    y=alt.Y('pct:Q', stack='zero', title='% of enrolled', scale=alt.Scale(domain=[0, 100])),
    color=alt.Color('bucket:N', scale=color_scale, title=None),
    order='bucket_order:Q',
    tooltip=[alt.Tooltip('snapshot_label:N', title='Snapshot'), alt.Tooltip('bucket:N', title='Segment'), alt.Tooltip('pct:Q', title='%', format='.1f')],
)

line = alt.Chart(sub_df).mark_line(point=alt.OverlayMarkDef(color='black'), color='black').encode(
    x=alt.X('snapshot_label:N', sort=snapshot_order),
    y=alt.Y('red_pct:Q'),
    tooltip=[alt.Tooltip('snapshot_label:N', title='Snapshot'), alt.Tooltip('red_pct:Q', title='S1+S2+S3 %', format='.1f')],
)

st.altair_chart((bar + line).properties(height=420), use_container_width=True)

st.dataframe(
    sub_df[['snapshot_label', 'S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'enrolled']],
    use_container_width=True, hide_index=True,
)

st.subheader('4. Save your progress')
combined_out = combined.copy()
combined_out['snapshot_date'] = combined_out['snapshot_date'].dt.strftime('%Y-%m-%d')
csv_buf = io.StringIO()
combined_out.to_csv(csv_buf, index=False)
st.download_button(
    '⬇ Download updated trend history (CSV)',
    data=csv_buf.getvalue(),
    file_name='engagement_trend_history.csv',
    mime='text/csv',
)
st.caption(
    'This file contains only aggregate segment counts per snapshot (subject, date, '
    'S1-S7 totals, enrolled count) - no student-level data. Keep it somewhere you can '
    'find it and re-upload it next time to keep the trend going.'
)
