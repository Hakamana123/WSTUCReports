"""
WSUTC Student Engagement Report Builder
========================================
Streamlit app for generating weekly engagement reports for GEDU0016 and GEDU0017
(or any single Blackboard subject following the same export format).

Run with:
    streamlit run engagement_report_app.py

Then open the URL it prints (usually http://localhost:8501) in a browser.
"""
import io
import re
import xml.etree.ElementTree as ET
from datetime import date, datetime, timedelta

import streamlit as st
import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ===========================================================================
# CONSTANTS
# ===========================================================================
NS = 'urn:schemas-microsoft-com:office:spreadsheet'
P = f'{{{NS}}}'

DAY_TO_COL = {
    1: 5, 2: 7, 3: 8, 4: 9, 5: 10, 6: 11, 7: 13, 8: 15, 9: 17, 10: 19,
    11: 20, 12: 22, 13: 23, 14: 25, 15: 27, 16: 28, 17: 29, 18: 31, 19: 32,
    20: 33, 21: 34, 22: 35, 23: 37, 24: 38, 25: 39, 26: 41, 27: 42, 28: 43,
    29: 45, 30: 46, 31: 47,
}

WEEK1_START = date(2026, 3, 2)

EXCLUDE_SURNAMES = {'Curtin', 'Rouillon', 'Turro', 'Tyler', 'Wyborn', 'Wagstaffe', 'Pinkerton'}
USAGE_EXCLUDE_NAMES = {'Guest', 'Total'}

NAVY = '1F2F4E'
ACCENT = '2E5D9F'
LIGHT = 'EEF2F7'
ALT_ROW = 'F7F9FC'
WHITE = 'FFFFFF'
RED = 'C0392B'
ORANGE = 'D68910'
YELLOW = 'F1C40F'
GREEN = '27AE60'
BLUE = '2980B9'
PURPLE = '7D3C98'
SEG_COLOURS = {'S1': RED, 'S2': '8B4513', 'S3': ORANGE, 'S4': YELLOW,
               'S5': BLUE, 'S6': PURPLE, 'S7': GREEN}

GHOST_SEGS = {'S1', 'S2', 'S3'}

# ===========================================================================
# CLASS LIST PARSING
# ===========================================================================
def load_classlist(file_bytes):
    if file_bytes[:2] == b'PK':
        return _load_classlist_xlsx(file_bytes)
    return _load_classlist_xls(file_bytes)


def _load_classlist_xls(file_bytes):
    book = xlrd.open_workbook(file_contents=file_bytes)
    sheet = book.sheet_by_index(0)
    headers = [str(sheet.cell_value(6, c)).strip().lower() for c in range(sheet.ncols)]
    cmap = {h: i for i, h in enumerate(headers) if h}
    if 'student_code' not in cmap:
        raise ValueError("Class list header row not found at row 7. Check the file format.")
    cm = {
        'sid':    cmap['student_code'],
        'first':  cmap.get('first_name', 1),
        'last':   cmap.get('last_name', 2),
        'course': cmap.get('course', 5),
        'email':  cmap.get('email_address', 6),
    }
    subject_raw = str(sheet.cell_value(1, 0)).strip()
    subject_match = re.match(r'([A-Z]+\d+)', subject_raw)
    subject_code = subject_match.group(1) if subject_match else 'UNKNOWN'
    students = {}
    for r in range(7, sheet.nrows):
        sid = str(sheet.cell_value(r, cm['sid'])).strip()
        if not sid or sid == 'student_code':
            continue
        if sid.startswith('30') or sid.startswith('96'):
            continue
        last = str(sheet.cell_value(r, cm['last'])).strip()
        if last in EXCLUDE_SURNAMES:
            continue
        course_val = str(sheet.cell_value(r, cm['course'])).strip()
        students[sid] = {
            'sid': sid,
            'first': str(sheet.cell_value(r, cm['first'])).strip(),
            'last': last,
            'course': course_val,
            'course_code': course_val,
            'discipline_subject': '',
            'discipline_class': '',
            'email': str(sheet.cell_value(r, cm['email'])).strip(),
            'discipline_teacher': '',
            'subject_code': '',
        }
    return subject_code, students


def _load_classlist_xlsx(file_bytes):
    wb_cl = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb_cl.active
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            headers[str(val).strip()] = c
    required = {'Student ID', 'First Name', 'Last Name'}
    missing = required - set(headers.keys())
    if missing:
        raise ValueError(f"Enriched class list missing columns: {missing}")
    subject_code = 'UNKNOWN'
    subj_col_name = None
    for candidate in ('GEDU Subject', 'Subject Code'):
        if candidate in headers:
            subj_col_name = candidate
            break
    if subj_col_name:
        raw = str(ws.cell(2, headers[subj_col_name]).value or '').strip()
        m = re.match(r'([A-Z]+\d+)', raw)
        if m:
            subject_code = m.group(1)
    def cell_str(row, col_name, default_col=1):
        c = headers.get(col_name, default_col)
        v = ws.cell(row, c).value
        if v is None:
            return ''
        if isinstance(v, (int, float)):
            return str(int(v))
        return str(v).strip()
    students = {}
    for r in range(2, ws.max_row + 1):
        sid_raw = ws.cell(r, headers['Student ID']).value
        if sid_raw is None:
            continue
        sid = str(int(sid_raw)).strip() if isinstance(sid_raw, (int, float)) else str(sid_raw).strip()
        if not sid or not sid.isdigit():
            continue
        if sid.startswith('30') or sid.startswith('96'):
            continue
        last = cell_str(r, 'Last Name')
        if last in EXCLUDE_SURNAMES:
            continue
        course_code = cell_str(r, 'Course Code')
        students[sid] = {
            'sid': sid,
            'first': cell_str(r, 'First Name'),
            'last': last,
            'course': course_code,
            'course_code': course_code,
            'discipline_subject': cell_str(r, 'Discipline Subject'),
            'discipline_class': cell_str(r, 'Discipline Class'),
            'email': cell_str(r, 'Email'),
            'discipline_teacher': cell_str(r, 'Discipline Teacher'),
            'subject_code': cell_str(r, subj_col_name) if subj_col_name else '',
        }
    return subject_code, students

# ===========================================================================
# LOGIN REPORT PARSING
# ===========================================================================
def load_login_report(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    out = {}
    window_start = window_end = None
    for r in range(1, min(ws.max_row + 1, 60)):
        for c in range(1, min(ws.max_column + 1, 20)):
            v = ws.cell(r, c).value
            if isinstance(v, str) and 'between' in v and 'to' in v:
                m = re.search(r'between\s+(\d{1,2}/\d{1,2}/\d{4})\s+to\s+(\d{1,2}/\d{1,2}/\d{4})', v)
                if m:
                    try:
                        window_start = datetime.strptime(m.group(1), '%d/%m/%Y').date()
                        window_end = datetime.strptime(m.group(2), '%d/%m/%Y').date()
                    except ValueError:
                        pass
                if window_start: break
        if window_start: break
    nli_header_row = None
    li_header_row = None
    for r in range(1, ws.max_row + 1):
        v8 = ws.cell(r, 8).value
        v6 = ws.cell(r, 6).value
        if v8 == 'SURNAME' and nli_header_row is None:
            nli_header_row = r
        if v6 == 'SURNAME':
            li_header_row = r
    if nli_header_row is None or li_header_row is None:
        raise ValueError("Could not find login report section headers")
    for r in range(nli_header_row + 1, li_header_row):
        sid = ws.cell(r, 23).value
        if sid is None: continue
        sid = str(sid).strip()
        if not sid.isdigit(): continue
        days = ws.cell(r, 40).value
        last = ws.cell(r, 45).value
        total = ws.cell(r, 49).value
        never = (str(days).strip().upper() == 'NEVER')
        days_n = None if never else (int(days) if isinstance(days, (int, float)) else None)
        last_d = last if isinstance(last, (datetime, date)) else None
        out[sid] = {
            'days_since': days_n, 'last_login': last_d,
            'total_logins': int(total) if isinstance(total, (int, float)) else 0,
            'never': never, 'in_window': False,
        }
    for r in range(li_header_row + 1, ws.max_row + 1):
        sid = ws.cell(r, 22).value
        if sid is None: continue
        sid = str(sid).strip()
        if not sid.isdigit(): continue
        days = ws.cell(r, 39).value
        last = ws.cell(r, 44).value
        total = ws.cell(r, 48).value
        days_n = int(days) if isinstance(days, (int, float)) else None
        last_d = last if isinstance(last, (datetime, date)) else None
        out[sid] = {
            'days_since': days_n, 'last_login': last_d,
            'total_logins': int(total) if isinstance(total, (int, float)) else 0,
            'never': False, 'in_window': True,
        }
    return out, window_start, window_end

# ===========================================================================
# USAGE FILE PARSING
# ===========================================================================
def get_cells(row):
    cells = row.findall(f'{P}Cell'); result = {}; ci = 0
    for cell in cells:
        idx = cell.get(f'{P}Index')
        if idx: ci = int(idx) - 1
        d = cell.find(f'{P}Data')
        result[ci] = d.text if d is not None else ''
        ci += 1
    return result

def parse_usage_file(file_bytes):
    tree = ET.parse(io.BytesIO(file_bytes))
    rows = tree.getroot().findall(f'.//{P}Row')
    data = {}
    current_month = None
    for r in rows:
        c = get_cells(r)
        m = c.get(1)
        if isinstance(m, str) and re.match(r'^\d{4}-\d{2}$', m):
            current_month = m
            continue
        if current_month is None:
            continue
        name = (c.get(1) or '').strip()
        if not name:
            continue
        if (name.startswith('Total') or name.startswith('Overall')
                or name == 'Chart does not appear in Excel'):
            current_month = None
            continue
        if name in USAGE_EXCLUDE_NAMES or 'PreviewUser' in name:
            continue
        if '(' not in name or ')' not in name:
            continue
        sid = name[name.rfind('(')+1:name.rfind(')')].strip()
        if not sid.isdigit():
            continue
        if sid.startswith('30') or sid.startswith('96'):
            continue
        year_str, month_str = current_month.split('-')
        year = int(year_str); month = int(month_str)
        for day, col in DAY_TO_COL.items():
            v = c.get(col)
            if v is None or v == '':
                continue
            try:
                hits = int(float(v))
            except (ValueError, TypeError):
                continue
            key = (year, month, day)
            data.setdefault(key, {})[sid] = hits
    return data

def merge_usage_files(usage_file_list):
    merged = {}
    for file_bytes in usage_file_list:
        single = parse_usage_file(file_bytes)
        for date_key, student_hits in single.items():
            merged[date_key] = student_hits
    return merged

# ===========================================================================
# GRADE CENTRE PARSING
# ===========================================================================
def load_grade_centre(file_bytes):
    """
    Parse the Grade Centre export.

    gc_data structure:
        gc_data[sid][label] = {
            'original': str,   # raw value from the primary column ('' if blank)
            'resubmit': str,   # raw value from the resubmit column  ('' if blank)
        }

    The four derived metrics are computed later via the gc_* helper functions.
    """
    import csv as _csv
    text = file_bytes.decode('utf-16-le')
    if text and text[0] == '\ufeff':
        text = text[1:]
    reader = _csv.reader(text.splitlines(), delimiter='\t')
    headers_raw = next(reader)
    rows = list(reader)
    headers = [h.strip().strip('"') for h in headers_raw]

    sid_col = None
    for i, h in enumerate(headers):
        if h.lower() in ('username', 'student id'):
            sid_col = i
            break
    if sid_col is None:
        raise ValueError("Grade Centre: cannot find Username or Student ID column.")

    avail_col = None
    for i, h in enumerate(headers):
        if h.lower() == 'availability':
            avail_col = i
            break

    def short_name(h):
        return h.split('[')[0].strip() if '[' in h else h.strip()

    import re as _re

    primary_cols = []
    resubmit_cols = []
    for i, h in enumerate(headers):
        sn = short_name(h).lower()
        if not sn.startswith('assessment'):
            continue
        if 'resubmit' in sn or 'resubmission' in sn:
            resubmit_cols.append((i, short_name(h)))
        else:
            primary_cols.append((i, short_name(h)))

    collated_nums = set()
    for _, name in primary_cols:
        if 'collated' in name.lower() or 'total' in name.lower().split('assessment')[-1]:
            m = _re.search(r'Assessment\s+(\d+)', name, _re.IGNORECASE)
            if m:
                collated_nums.add(m.group(1))

    selected = []
    seen_nums = set()
    for i, name in primary_cols:
        m = _re.search(r'Assessment\s+(\d+)', name, _re.IGNORECASE)
        if not m: continue
        num = m.group(1)
        is_collated = ('collated' in name.lower() or 'total' in name.lower().split('assessment')[-1].split(':')[0])
        if num in collated_nums:
            if is_collated and num not in seen_nums:
                selected.append((i, name, num)); seen_nums.add(num)
        else:
            if num not in seen_nums:
                selected.append((i, name, num)); seen_nums.add(num)
    selected.sort(key=lambda x: int(x[2]))

    # Map assessment number -> resubmit column index
    resub_map = {}
    for i, name in resubmit_cols:
        m = _re.search(r'Assessment\s+(\d+)', name, _re.IGNORECASE)
        if m:
            num = m.group(1)
            if num in seen_nums:
                resub_map[num] = i

    gc_labels = [f'AS{num}' for _, _, num in selected]
    gc_col_indices = [i for i, _, _ in selected]
    gc_nums = [num for _, _, num in selected]

    gc_data = {}
    for row in rows:
        if len(row) <= sid_col: continue
        sid = row[sid_col].strip().strip('"')
        if not sid or not sid.isdigit(): continue
        if sid.startswith('30') or sid.startswith('96'): continue
        if avail_col is not None and row[avail_col].strip().strip('"').lower() != 'yes': continue
        if 'PreviewUser' in (row[0] if row else ''): continue

        student_gc = {}
        for label, col_idx, num in zip(gc_labels, gc_col_indices, gc_nums):
            # Raw original value
            original_raw = row[col_idx].strip().strip('"') if col_idx < len(row) else ''

            # Raw resubmit value
            resub_idx = resub_map.get(num)
            resubmit_raw = ''
            if resub_idx is not None and resub_idx < len(row):
                resubmit_raw = row[resub_idx].strip().strip('"')

            student_gc[label] = {
                'original': original_raw,
                'resubmit': resubmit_raw,
            }
        gc_data[sid] = student_gc
    return gc_data, gc_labels


# ===========================================================================
# GRADE CENTRE METRIC HELPERS
# ===========================================================================

def gc_submitted(entry):
    """Submission: original column has any non-blank value."""
    return bool(entry.get('original', ''))


def gc_passed(entry):
    """
    Pass: original = Satisfactory, OR
          original failed/needs grading AND resubmit = Satisfactory (or numeric >= 50).
    """
    original = entry.get('original', '').lower()
    resubmit = entry.get('resubmit', '').lower()

    if 'satisf' in original and 'unsatisf' not in original:
        return True  # original pass

    # Original was a fail or needs grading — check resubmit
    if original and ('unsatisf' in original or 'needs' in original):
        if 'satisf' in resubmit and 'unsatisf' not in resubmit:
            return True
        try:
            return float(resubmit) >= 50
        except (ValueError, TypeError):
            pass

    return False


def gc_needed_resubmit(entry):
    """Needed to resubmit: original = Unsatisfactory or Needs Grading/Marking."""
    original = entry.get('original', '').lower()
    return bool(original) and ('unsatisf' in original or 'needs grading' in original or 'needs marking' in original)


def gc_resubmitted(entry):
    """
    Resubmission: needed to resubmit AND actually submitted something in the resubmit column.
    """
    return gc_needed_resubmit(entry) and bool(entry.get('resubmit', ''))


def gc_resubmit_passed(entry):
    """
    Resubmission pass: needed to resubmit AND resubmit outcome = Satisfactory (or score >= 50).
    """
    if not gc_needed_resubmit(entry):
        return False
    resubmit = entry.get('resubmit', '').lower()
    if 'satisf' in resubmit and 'unsatisf' not in resubmit:
        return True
    try:
        return float(resubmit) >= 50
    except (ValueError, TypeError):
        return False


def gc_final_outcome(entry):
    """
    Final outcome for cell colouring (where they ended up):
      'Satisfactory'   — passed (original or via resubmit)
      'Unsatisfactory' — failed and either did not resubmit or resubmit also failed
      'Needs Grading'  — submitted but outcome not yet determined
      'No Submission'  — nothing submitted
    """
    if not gc_submitted(entry):
        return 'No Submission'

    if gc_passed(entry):
        return 'Satisfactory'

    resubmit = entry.get('resubmit', '').lower()
    original = entry.get('original', '').lower()

    # If resubmit column has something but didn't pass
    if resubmit:
        if 'needs' in resubmit:
            return 'Needs Grading'
        return 'Unsatisfactory'

    # No resubmit yet — judge by original
    if 'needs' in original:
        return 'Needs Grading'
    if 'unsatisf' in original:
        return 'Unsatisfactory'

    # Satisfactory already handled above
    return 'Needs Grading'


def gc_display_value(entry):
    """
    Human-readable string for display in per-student cells.
    Shows final outcome label so readers see where the student ended up.
    """
    outcome = gc_final_outcome(entry)
    if outcome == 'No Submission':
        return 'No Submission'
    if outcome == 'Satisfactory':
        original = entry.get('original', '').lower()
        if 'satisf' in original and 'unsatisf' not in original:
            return 'Satisfactory'
        return 'Satisfactory (via Resubmit)'
    if outcome == 'Unsatisfactory':
        resubmit = entry.get('resubmit', '')
        if resubmit:
            return f'Resub Fail'
        return 'Unsatisfactory'
    return 'Needs Grading'


# ===========================================================================
# WEEK DETECTION
# ===========================================================================
def detect_current_week(merged_usage, override_latest=None):
    if not merged_usage: return None, 0, None
    all_dates = [date(y, m, d) for (y, m, d) in merged_usage.keys()]
    if override_latest:
        eligible = [d for d in all_dates if d <= override_latest]
        latest = max(eligible) if eligible else override_latest
    else:
        latest = max(all_dates)
    if latest < WEEK1_START: return None, 0, latest
    week_num = ((latest - WEEK1_START).days // 7) + 1
    week_start = WEEK1_START + timedelta(days=(week_num - 1) * 7)
    days_in = (latest - week_start).days + 1
    return week_num, days_in, latest

def bucket_by_week(merged_usage, students, max_week, max_date=None):
    hits = {sid: {f'w{i}': 0 for i in range(1, max_week + 1)} for sid in students}
    for (y, m, d), student_hits in merged_usage.items():
        dt = date(y, m, d)
        if dt < WEEK1_START: continue
        if max_date and dt > max_date: continue
        week_num = ((dt - WEEK1_START).days // 7) + 1
        if week_num > max_week: continue
        wkey = f'w{week_num}'
        for sid, h in student_hits.items():
            if sid in hits: hits[sid][wkey] += h
    return hits

def week_date_range(week_num):
    start = WEEK1_START + timedelta(days=(week_num - 1) * 7)
    return start, start + timedelta(days=6)

# ===========================================================================
# SEGMENTATION
# ===========================================================================
def classify(students, login, hits, current_week, prev_days, curr_days):
    seg = {}
    week_keys = [f'w{i}' for i in range(1, current_week + 1)]
    curr_key = f'w{current_week}'
    prev_key = f'w{current_week - 1}' if current_week > 1 else None
    s2_threshold = (current_week - 1) * 7 + curr_days
    s3_low = s2_threshold - 7
    s3_high = s2_threshold - 1
    for sid in students:
        h = hits[sid]; l = login.get(sid)
        zero_usage = all(h[k] == 0 for k in week_keys)
        days_since = None; in_window = False; never_flag = False
        if l is not None:
            days_since = l['days_since']; in_window = l['in_window']; never_flag = l['never']
        if zero_usage and (never_flag or l is None): seg[sid] = 'S1'; continue
        if zero_usage and days_since is not None and days_since >= s2_threshold: seg[sid] = 'S2'; continue
        if zero_usage: seg[sid] = 'S2'; continue
        if (days_since is not None and s3_low <= days_since <= s3_high and not in_window): seg[sid] = 'S3'; continue
        if h[curr_key] == 0:
            if any(h[k] > 0 for k in week_keys[:-1]): seg[sid] = 'S4'; continue
        if prev_key:
            prev = h[prev_key]; curr = h[curr_key]
            if prev == 0 and curr > 0: seg[sid] = 'S5'; continue
            if prev > 0 and curr > 0:
                seg[sid] = 'S6' if (curr / curr_days) < (prev / prev_days) * 0.5 else 'S7'; continue
        seg[sid] = 'S1'
    return seg, s2_threshold, (s3_low, s3_high)

# ===========================================================================
# WORKBOOK BUILDING HELPERS
# ===========================================================================
def thin_border():
    side = Side(style='thin', color='D5DBDB')
    return Border(left=side, right=side, top=side, bottom=side)

def write_tab_header(ws, title, subtitle, description, n_cols, seg_code=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1, title); c.font = Font(name='Arial', size=14, bold=True, color=WHITE)
    c.fill = PatternFill('solid', start_color=NAVY); c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 26
    seg_colour = SEG_COLOURS.get(seg_code, ACCENT) if seg_code else ACCENT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws.cell(2, 1, subtitle or ''); c.font = Font(name='Arial', size=11, bold=True, color=WHITE)
    c.fill = PatternFill('solid', start_color=seg_colour); c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 20
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    c = ws.cell(3, 1, description); c.font = Font(name='Arial', size=10, italic=True, color='2C3E50')
    c.fill = PatternFill('solid', start_color=LIGHT); c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    ws.row_dimensions[3].height = 36; ws.row_dimensions[4].height = 6

def write_col_headers(ws, headers, row=5):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h); c.font = Font(name='Arial', size=10, bold=True, color=WHITE)
        c.fill = PatternFill('solid', start_color=ACCENT); c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 30

def write_data_rows(ws, data_rows, start_row=6):
    for ri, row in enumerate(data_rows):
        excel_row = start_row + ri
        fill = PatternFill('solid', start_color=ALT_ROW) if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(excel_row, ci, val); c.font = Font(name='Arial', size=10)
            if fill: c.fill = fill
            c.alignment = Alignment(horizontal='left' if ci <= 5 else 'center', vertical='center')
            c.border = thin_border()

def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def fmt_login(l):
    if l is None: return ('NO LOGIN RECORD', '-', 0)
    if l['never']: return ('NEVER', 'NEVER', l['total_logins'] or 0)
    days = l['days_since']; last = l['last_login']
    last_str = last.strftime('%Y-%m-%d') if isinstance(last, (datetime, date)) else (str(last) if last else '-')
    return (days if days is not None else '-', last_str, l['total_logins'] or 0)

# ===========================================================================
# GRADE CENTRE RATE TABLE HELPERS
# ===========================================================================

STATUS_COLS = ['Satisfactory', 'Unsatisfactory', 'Needs Grading', 'No Submission']


def _write_rate_section(ws, start_row, title, note, col_headers, gc_labels, students,
                        seg, enrolled_fn, count_fn, n_cols):
    """
    Generic rate table writer.

    enrolled_fn(sid) -> bool   counts who is in the denominator for this metric
    count_fn(sid, label) -> bool   counts who hits the numerator
    Returns next available row.
    """
    ghost_sids = {sid for sid, s in seg.items() if s in GHOST_SEGS}
    enrolled_all = len(students)
    non_ghost_enrolled = sum(1 for sid in students if sid not in ghost_sids)

    # Section header
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=n_cols)
    c = ws.cell(start_row, 1, title)
    c.font = Font(name='Arial', size=11, bold=True, color=WHITE)
    c.fill = PatternFill('solid', start_color=NAVY)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[start_row].height = 22
    start_row += 1

    # Note row
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=n_cols)
    c = ws.cell(start_row, 1, note)
    c.font = Font(name='Arial', size=9, italic=True, color='2C3E50')
    c.fill = PatternFill('solid', start_color=LIGHT)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    ws.row_dimensions[start_row].height = 18
    start_row += 1

    # Column headers
    for ci, h in enumerate(col_headers, 1):
        c = ws.cell(start_row, ci, h)
        c.font = Font(name='Arial', size=10, bold=True, color=WHITE)
        c.fill = PatternFill('solid', start_color=ACCENT)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[start_row].height = 30
    start_row += 1

    # Data rows — one per assessment
    for ri, label in enumerate(gc_labels):
        denom_all   = sum(1 for sid in students if enrolled_fn(sid, label, 'all'))
        denom_ng    = sum(1 for sid in students if sid not in ghost_sids and enrolled_fn(sid, label, 'non_ghost'))
        n_all       = sum(1 for sid in students if enrolled_fn(sid, label, 'all')      and count_fn(sid, label))
        n_ng        = sum(1 for sid in students if sid not in ghost_sids
                          and enrolled_fn(sid, label, 'non_ghost') and count_fn(sid, label))

        rate_all = n_all / denom_all if denom_all else 0
        rate_ng  = n_ng  / denom_ng  if denom_ng  else 0

        fill = PatternFill('solid', start_color=ALT_ROW) if ri % 2 == 0 else None
        row_vals = [label.replace('AS', 'Assessment '), n_all, rate_all, n_ng, rate_ng]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(start_row, ci, val)
            c.font = Font(name='Arial', size=10)
            if fill:
                c.fill = fill
            c.border = thin_border()
            if ci in (3, 5):
                c.number_format = '0.0%'
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 1:
                c.alignment = Alignment(horizontal='left', vertical='center')
            else:
                c.alignment = Alignment(horizontal='center', vertical='center')
        start_row += 1

    # Spacer
    ws.row_dimensions[start_row].height = 8
    start_row += 1
    return start_row


def _write_all_rate_tables(ws, start_row, gc_data, gc_labels, students, seg, n_cols):
    """
    Write four rate tables:
      1. Submission rate      — submitted / all enrolled
      2. Pass rate            — passed / all enrolled
      3. Resubmission rate    — resubmitted / those who needed to resubmit
      4. Resubmission pass rate — resubmit passed / those who needed to resubmit
    """
    ghost_sids = {sid for sid, s in seg.items() if s in GHOST_SEGS}
    enrolled = len(students)
    non_ghost_enrolled = sum(1 for sid in students if sid not in ghost_sids)
    suffix = (
        f'Ghosts = S1+S2+S3 ({len(ghost_sids)} students).  '
        f'Enrolled: {enrolled}  |  Excl. Ghosts: {non_ghost_enrolled}'
    )

    def all_enrolled(sid, label, scope):
        return True

    def non_ghost_enrolled_fn(sid, label, scope):
        return sid not in ghost_sids

    def needed_resub(sid, label, scope):
        entry = gc_data.get(sid, {}).get(label, {})
        result = gc_needed_resubmit(entry)
        if scope == 'non_ghost':
            return result and sid not in ghost_sids
        return result

    # 1. Submission rate
    start_row = _write_rate_section(
        ws, start_row,
        title='Assessment Submission Rates',
        note='Submitted = any value in the original submission column.  ' + suffix,
        col_headers=[
            'Assessment', 'Submitted (n)', 'Rate % (all enrolled)',
            'Submitted excl. Ghosts (n)', 'Rate % (excl. Ghosts)',
        ],
        gc_labels=gc_labels, students=students, seg=seg,
        enrolled_fn=lambda sid, label, scope: True,
        count_fn=lambda sid, label: gc_submitted(gc_data.get(sid, {}).get(label, {})),
        n_cols=n_cols,
    )

    # 2. Pass rate
    start_row = _write_rate_section(
        ws, start_row,
        title='Assessment Pass Rates',
        note='Pass = Satisfactory on original submission, or Satisfactory on resubmission.  ' + suffix,
        col_headers=[
            'Assessment', 'Passed (n)', 'Pass Rate % (all enrolled)',
            'Passed excl. Ghosts (n)', 'Pass Rate % (excl. Ghosts)',
        ],
        gc_labels=gc_labels, students=students, seg=seg,
        enrolled_fn=lambda sid, label, scope: True,
        count_fn=lambda sid, label: gc_passed(gc_data.get(sid, {}).get(label, {})),
        n_cols=n_cols,
    )

    # 3. Resubmission rate  (denominator = those who needed to resubmit)
    start_row = _write_rate_section(
        ws, start_row,
        title='Assessment Resubmission Rates',
        note=(
            'Denominator = students who needed to resubmit (original = Unsatisfactory or Needs Grading).  '
            'Numerator = those who actually resubmitted (regardless of outcome).  ' + suffix
        ),
        col_headers=[
            'Assessment', 'Resubmitted (n)', 'Resub Rate % (of those needing resub)',
            'Resubmitted excl. Ghosts (n)', 'Resub Rate % (excl. Ghosts)',
        ],
        gc_labels=gc_labels, students=students, seg=seg,
        enrolled_fn=lambda sid, label, scope: gc_needed_resubmit(gc_data.get(sid, {}).get(label, {})),
        count_fn=lambda sid, label: gc_resubmitted(gc_data.get(sid, {}).get(label, {})),
        n_cols=n_cols,
    )

    # 4. Resubmission pass rate  (denominator = those who needed to resubmit)
    start_row = _write_rate_section(
        ws, start_row,
        title='Assessment Resubmission Pass Rates',
        note=(
            'Denominator = students who needed to resubmit (original = Unsatisfactory or Needs Grading).  '
            'Numerator = those who passed on resubmission.  ' + suffix
        ),
        col_headers=[
            'Assessment', 'Resub Passed (n)', 'Resub Pass Rate % (of those needing resub)',
            'Resub Passed excl. Ghosts (n)', 'Resub Pass Rate % (excl. Ghosts)',
        ],
        gc_labels=gc_labels, students=students, seg=seg,
        enrolled_fn=lambda sid, label, scope: gc_needed_resubmit(gc_data.get(sid, {}).get(label, {})),
        count_fn=lambda sid, label: gc_resubmit_passed(gc_data.get(sid, {}).get(label, {})),
        n_cols=n_cols,
    )

    return start_row


def _write_assessment_detail_sheet(wb, gc_data, gc_labels, groups, group_label_col, current_week, is_partial):
    """
    Create a new 'Assessment Detail' sheet with per-class breakdown per assessment.
    Layout: one section per assessment, rows = groups (classes), columns = final-outcome status categories.
    """
    ws = wb.create_sheet('Assessment Detail')
    partial_note = ' (PARTIAL)' if is_partial else ''
    n_cols = len(STATUS_COLS) + 2  # group col + status cols + total

    write_tab_header(
        ws,
        f'Assessment Detail — W{current_week}{partial_note}',
        f'Submission status breakdown by {group_label_col.lower()} for each assessment',
        'One section per assessment. Columns show final outcome (where the student ended up).',
        n_cols,
    )

    group_order = sorted(groups.keys(), key=lambda k: (-len(groups[k]['sids']), str(k)))
    current_row = 5

    for ai, label in enumerate(gc_labels):
        # Assessment section header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=n_cols)
        c = ws.cell(current_row, 1, label.replace('AS', 'Assessment '))
        c.font = Font(name='Arial', size=11, bold=True, color=WHITE)
        c.fill = PatternFill('solid', start_color=NAVY)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Column headers
        as_headers = [group_label_col] + STATUS_COLS + ['TOTAL']
        for ci, h in enumerate(as_headers, 1):
            c = ws.cell(current_row, ci, h)
            c.font = Font(name='Arial', size=10, bold=True, color=WHITE)
            c.fill = PatternFill('solid', start_color=ACCENT)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin_border()
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Data rows
        grand_totals = {s: 0 for s in STATUS_COLS}
        grand_total_all = 0

        for ri_idx, gk in enumerate(group_order):
            sids = groups[gk]['sids']
            bucket = {s: 0 for s in STATUS_COLS}
            for sid in sids:
                entry = gc_data.get(sid, {}).get(label, {})
                outcome = gc_final_outcome(entry)
                bucket[outcome] = bucket.get(outcome, 0) + 1
            row_total = sum(bucket.values())
            row_data = [gk] + [bucket.get(s, 0) for s in STATUS_COLS] + [row_total]

            fill = PatternFill('solid', start_color=ALT_ROW) if ri_idx % 2 == 0 else None
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(current_row, ci, val)
                c.font = Font(name='Arial', size=10)
                if fill:
                    c.fill = fill
                c.alignment = Alignment(
                    horizontal='center' if ci > 1 else 'left', vertical='center'
                )
                c.border = thin_border()
            current_row += 1

            for s in STATUS_COLS:
                grand_totals[s] = grand_totals.get(s, 0) + bucket.get(s, 0)
            grand_total_all += row_total

        # Totals row
        total_row_data = ['TOTAL'] + [grand_totals.get(s, 0) for s in STATUS_COLS] + [grand_total_all]
        for ci, val in enumerate(total_row_data, 1):
            c = ws.cell(current_row, ci, val)
            c.font = Font(name='Arial', size=10, bold=True)
            c.fill = PatternFill('solid', start_color=LIGHT)
            c.border = thin_border()
            c.alignment = Alignment(
                horizontal='center' if ci > 1 else 'left', vertical='center'
            )
        current_row += 2  # spacer between assessments

    # Column widths
    widths = [28] + [14] * len(STATUS_COLS) + [10]
    autosize(ws, widths)
    ws.freeze_panes = 'A5'
    return ws


def _write_class_index_sheet(wb, title_prefix, groups, group_label_col, students, seg,
                              gc_data, gc_labels, group_sheet_names, current_week, is_partial):
    ws = wb.create_sheet('Class Index')
    partial_note = ' (PARTIAL)' if is_partial else ''
    gc_active = gc_data is not None and gc_labels is not None and len(gc_labels) > 0

    n_fixed = 6
    n_cols = n_fixed + (len(gc_labels) if gc_active else 0)

    write_tab_header(
        ws,
        f'{title_prefix} — Class Index{partial_note}',
        f'{len(groups)} classes  •  Click any class to jump to its sheet',
        f'Active = S5+S6+S7.  At Risk = S4.  '
        + ('AS# Sub = submitted / enrolled.  Green = 100%, Orange ≥ 50%, Red < 50%.'
           if gc_active else ''),
        n_cols,
    )

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)
    c = ws.cell(4, 1,
        f'Click a class name to navigate directly. '
        f'Each class sheet has a "\u2190 Back to Index" link.  \u2022  '
        + ('Assessment columns show submitted\u202f/\u202fenrolled'
           if gc_active else ''))
    c.font = Font(name='Arial', size=9, italic=True, color='2C3E50')
    c.fill = PatternFill('solid', start_color=LIGHT)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[4].height = 14

    col_headers = ['#', group_label_col, 'Teacher', 'Enrolled', 'Active', 'At Risk']
    if gc_active:
        col_headers += [f'{lbl} Sub' for lbl in gc_labels]
    for ci, h in enumerate(col_headers, 1):
        c = ws.cell(5, ci, h)
        c.font = Font(name='Arial', size=10, bold=True, color=WHITE)
        c.fill = PatternFill('solid', start_color=NAVY)
        c.alignment = Alignment(horizontal='center' if ci != 2 else 'left',
                                vertical='center', wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[5].height = 30

    fill_green  = PatternFill('solid', start_color='D5F5E3')
    fill_orange = PatternFill('solid', start_color='FDEBD0')
    fill_red    = PatternFill('solid', start_color='FADBD8')
    font_green  = Font(name='Arial', size=10, color='1A7A3A', bold=True)
    font_orange = Font(name='Arial', size=10, color=ORANGE, bold=True)
    font_red    = Font(name='Arial', size=10, color=RED, bold=True)

    group_order = sorted(groups.keys(), key=lambda k: (-len(groups[k]['sids']), str(k)))

    for ri, gk in enumerate(group_order):
        excel_row = 6 + ri
        info = groups[gk]
        sids = info['sids']
        teacher = info.get('teacher', '')
        enrolled = len(sids)
        active   = sum(1 for sid in sids if seg.get(sid) in ('S5', 'S6', 'S7'))
        at_risk  = sum(1 for sid in sids if seg.get(sid) == 'S4')

        row_fill = PatternFill('solid', start_color=ALT_ROW) if ri % 2 == 0 else None

        fixed_vals = [ri + 1, gk, teacher, enrolled, active, at_risk]
        for ci, val in enumerate(fixed_vals, 1):
            c = ws.cell(excel_row, ci, val)
            c.font = Font(name='Arial', size=10)
            if row_fill:
                c.fill = row_fill
            c.alignment = Alignment(
                horizontal='left' if ci == 2 else 'center', vertical='center'
            )
            c.border = thin_border()

        if at_risk > 0:
            ws.cell(excel_row, 6).font = Font(name='Arial', size=10, color=RED, bold=True)

        sn = group_sheet_names.get(gk)
        if sn:
            cell = ws.cell(excel_row, 2)
            cell.hyperlink = f"#'{sn}'!A1"
            cell.font = Font(name='Arial', size=10, color='2980B9', underline='single')

        if gc_active:
            for ai, label in enumerate(gc_labels):
                col = n_fixed + 1 + ai
                submitted = sum(
                    1 for sid in sids
                    if gc_submitted(gc_data.get(sid, {}).get(label, {}))
                )
                frac_str = f'{submitted}/{enrolled}'
                rate = submitted / enrolled if enrolled else 0

                c = ws.cell(excel_row, col, frac_str)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = thin_border()
                if rate == 1.0:
                    c.fill = fill_green;  c.font = font_green
                elif rate >= 0.5:
                    c.fill = fill_orange; c.font = font_orange
                else:
                    c.fill = fill_red;   c.font = font_red

    totals_row = 6 + len(group_order)
    total_enrolled = sum(len(groups[gk]['sids']) for gk in group_order)
    total_active   = sum(
        sum(1 for sid in groups[gk]['sids'] if seg.get(sid) in ('S5', 'S6', 'S7'))
        for gk in group_order
    )
    total_at_risk  = sum(
        sum(1 for sid in groups[gk]['sids'] if seg.get(sid) == 'S4')
        for gk in group_order
    )
    total_vals = ['', 'TOTAL', '', total_enrolled, total_active, total_at_risk]
    if gc_active:
        for label in gc_labels:
            total_sub = sum(
                1 for gk in group_order for sid in groups[gk]['sids']
                if gc_submitted(gc_data.get(sid, {}).get(label, {}))
            )
            total_vals.append(f'{total_sub}/{total_enrolled}')
    for ci, val in enumerate(total_vals, 1):
        c = ws.cell(totals_row, ci, val)
        c.font = Font(name='Arial', size=10, bold=True)
        c.fill = PatternFill('solid', start_color=LIGHT)
        c.border = thin_border()
        c.alignment = Alignment(horizontal='center' if ci != 2 else 'left', vertical='center')

    widths = [4, 36, 22, 10, 10, 10] + ([9] * len(gc_labels) if gc_active else [])
    autosize(ws, widths)
    ws.freeze_panes = 'C6'

    for gk in group_order:
        sn = group_sheet_names.get(gk)
        if sn and sn in [s.title for s in wb.worksheets]:
            ws_g = wb[sn]
            back_cell = ws_g.cell(3, 1)
            existing_val = back_cell.value or ''
            back_cell.value = '\u2190 Back to Index      ' + existing_val
            back_cell.hyperlink = "#'Class Index'!A1"
            back_cell.font = Font(name='Arial', size=10, italic=True,
                                  color='2980B9', underline='single')

    return ws


# ===========================================================================
# MAIN ENGAGEMENT WORKBOOK
# ===========================================================================
def build_workbook(subject_code, students, login, hits, seg, current_week, prev_days, curr_days,
                   is_partial, latest_date, login_window, s2_threshold, s3_range):
    wb = Workbook(); wb.remove(wb.active)
    enrolled = len(students)
    counts = {f'S{i}': 0 for i in range(1, 8)}
    for s in seg.values(): counts[s] = counts.get(s, 0) + 1
    week_keys = [f'w{i}' for i in range(1, current_week + 1)]
    week_labels = [f'W{i}' for i in range(1, current_week + 1)]
    curr_key = f'w{current_week}'
    prev_key = f'w{current_week - 1}' if current_week > 1 else None
    curr_start, curr_end = week_date_range(current_week)
    if is_partial:
        curr_label = f'W{current_week} ({curr_start.strftime("%b %-d")}-{latest_date.strftime("%-d")}, {curr_days}d partial)'
    else:
        curr_label = f'W{current_week} ({curr_start.strftime("%b %-d")}-{curr_end.strftime("%-d")}, 7d)'
    if prev_key:
        prev_start, prev_end = week_date_range(current_week - 1)
        prev_label = f'W{current_week - 1} ({prev_start.strftime("%b %-d")}-{prev_end.strftime("%-d")}, 7d)'
    else:
        prev_label = '(none)'
    week_descriptions = ', '.join(f'W{i}={week_date_range(i)[0].strftime("%b %-d")}-{week_date_range(i)[1].strftime("%-d")}' for i in range(1, current_week + 1))

    seg_descriptions = {
        'S1': f'Never engaged: never logged in at all (or no login record) AND zero hits across W1 through W{current_week}.',
        'S2': 'Pre-teaching ghosts: last login was before Mar 2 (i.e. logged in during pre-teaching but never returned) AND zero hits all weeks.',
        'S3': 'W1 early drop-offs: last login fell in W1 and they have not returned in the current login window.',
        'S4': f'Active then absent in W{current_week}: had hits in a previous week but zero in W{current_week} to date. Split into "Just Dropped" (W{current_week-1}>0) and "Long Silent" (W1-W{current_week-2} active, W{current_week-1}+W{current_week} zero).' if current_week >= 3 else f'Active then absent in W{current_week}.',
        'S5': f'Late arrivals + W{current_week-1} returners: zero hits in W{current_week-1} but appearing in W{current_week}.' if current_week > 1 else 'Late arrivals: appearing in W1.',
        'S6': f'Fading engagers: active both weeks but daily-average hit rate fell 50%+ from W{current_week-1} to W{current_week}.' if current_week > 1 else 'Fading engagers (n/a in W1).',
        'S7': 'Sustained participants: active both weeks with daily-average rate held within 50%.' if current_week > 1 else 'Sustained participants (n/a in W1).',
    }
    seg_meta = [('S1','Never Engaged','Critical'),('S2','Pre-Teaching Ghosts','Critical'),('S3','W1 Early Drop-Offs','High Risk'),
                ('S4',f'Active then W{current_week} Absent','Watch (partial)' if is_partial else 'Watch'),
                ('S5','Late Arrivals + Returners','Mixed'),('S6','Fading Engagers','High Risk'),('S7','Sustained Participants','Healthy')]

    # Summary sheet
    ws = wb.create_sheet('Summary')
    write_tab_header(ws, f'{subject_code} Engagement Report — {curr_label}',
        f'Login window {login_window}  •  Enrolled {enrolled}  •  Latest data: {latest_date.strftime("%b %-d %Y")}',
        ('Partial-week run. S4 inflated, S7 understated by timing artefact.' if is_partial else 'Full-week run. Standard segmentation.'), 6)
    write_col_headers(ws, ['Segment','Label','Count','% of Enrolled','Status','Description'], row=5)
    rows = [[code, label, counts[code], counts[code]/enrolled, status, seg_descriptions[code]] for code, label, status in seg_meta]
    write_data_rows(ws, rows, start_row=6)
    for ri in range(6, 6+len(rows)): ws.cell(ri, 4).number_format = '0.0%'
    total_row = 6 + len(rows)
    ws.cell(total_row,1,'TOTAL').font = Font(name='Arial',size=10,bold=True)
    ws.cell(total_row,3,sum(counts.values())).font = Font(name='Arial',size=10,bold=True)
    ws.cell(total_row,4,sum(counts.values())/enrolled).number_format = '0.0%'
    ws.cell(total_row,4).font = Font(name='Arial',size=10,bold=True)
    for ci in range(1,7): ws.cell(total_row,ci).fill = PatternFill('solid',start_color=LIGHT); ws.cell(total_row,ci).border = thin_border()
    s3_or_s4_eligible = 0
    for sid in students:
        h = hits[sid]; l = login.get(sid)
        ds = l['days_since'] if l else None; iw = l['in_window'] if l else False
        s3e = (ds is not None and s3_range[0] <= ds <= s3_range[1] and not iw)
        s4e = (h[curr_key] == 0 and any(h[k] > 0 for k in week_keys[:-1]))
        if s3e and s4e: s3_or_s4_eligible += 1
    note_row = total_row + 2
    notes = [f'Enrolled (after exclusions): {enrolled}', f'Login window: {login_window}', f'Teaching weeks: {week_descriptions}',
        f'Comparison pair: {prev_label} vs {curr_label}. Daily averages normalised by actual day count.',
        (f'PARTIAL WEEK WARNING: W{current_week} has {curr_days} days of data. S4 inflated by timing. S7 understated. S5 includes both true late arrivals and W{current_week-1} returners. Do NOT use S4 list for outreach until W{current_week} closes.' if is_partial else f'Full week W{current_week}: standard run.'),
        f'Days-since thresholds: S2 ≥ {s2_threshold} days (last login on or before Mar 1); S3 in {s3_range[0]}-{s3_range[1]} days (last login Mar 2-8).',
        f'S3 / S4 dual-eligible students: {s3_or_s4_eligible}.', f'Students in class list but missing from login report: {sum(1 for s in students if s not in login)}.', 'Leaderboard ranking: hits per enrolled student (weighted engagement).']
    for i, n in enumerate(notes):
        c = ws.cell(note_row+i, 1, n); ws.merge_cells(start_row=note_row+i, start_column=1, end_row=note_row+i, end_column=6)
        c.font = Font(name='Arial', size=9, italic=(i==4), color=(RED if i==4 else '2C3E50'), bold=(i==4))
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[note_row+i].height = 32 if i==4 else 16
    autosize(ws, [10,28,10,14,18,70]); ws.freeze_panes = 'A6'

    def make_seg_tab(code, label):
        sids = sorted([sid for sid, s in seg.items() if s == code], key=lambda x: (students[x]['last'].lower(), students[x]['first'].lower()))
        ws = wb.create_sheet(f'{code} {label}'[:31])
        return ws, sids

    # S1
    ws, sids = make_seg_tab('S1','Never Engaged')
    write_tab_header(ws,'S1 — Never Engaged',f'{len(sids)} students with zero hits across all teaching weeks',seg_descriptions['S1'],10,'S1')
    write_col_headers(ws,['Surname','First Name','Student ID','Course','Disc. Class','Disc. Teacher','Email','Total Logins','Last Login','Action Required'],row=5)
    rows = []
    for sid in sids:
        st = students[sid]; days, last_str, total = fmt_login(login.get(sid))
        rows.append([st['last'],st['first'],sid,st['course'],st.get('discipline_class',''),st.get('discipline_teacher',''),st['email'],total,last_str,'Initial outreach — verify enrolment intent'])
    write_data_rows(ws, rows, start_row=6); autosize(ws,[22,18,12,10,10,18,38,12,14,38]); ws.freeze_panes = 'A6'

    # S2
    ws, sids = make_seg_tab('S2','Pre-Teaching Ghosts')
    write_tab_header(ws,'S2 — Pre-Teaching Ghosts',f'{len(sids)} students with zero hits all weeks AND last login pre-Mar 2 or NEVER',seg_descriptions['S2'],11,'S2')
    write_col_headers(ws,['Surname','First Name','Student ID','Course','Disc. Class','Disc. Teacher','Email','Total Logins','Days Since','Last Login','Action Required'],row=5)
    rows = []
    for sid in sids:
        st = students[sid]; days, last_str, total = fmt_login(login.get(sid))
        action = 'Single login only — likely orientation visit; escalate' if total == 1 else 'Pre-teaching login then disengaged — urgent outreach'
        rows.append([st['last'],st['first'],sid,st['course'],st.get('discipline_class',''),st.get('discipline_teacher',''),st['email'],total,days,last_str,action])
    write_data_rows(ws, rows, start_row=6); autosize(ws,[22,18,12,10,10,18,38,12,12,14,50]); ws.freeze_panes = 'A6'

    # S3
    ws, sids = make_seg_tab('S3','W1 Drop-Offs')
    write_tab_header(ws,'S3 — W1 Early Drop-Offs',f'{len(sids)} students whose last login fell in W1 (Mar 2-8)',seg_descriptions['S3'],12,'S3')
    write_col_headers(ws,['Surname','First Name','Student ID','Course','Disc. Class','Disc. Teacher','Email','Total Logins','Days Since','Last Login','Risk Level','Action Required'],row=5)
    rows = []
    for sid in sids:
        st = students[sid]; days, last_str, total = fmt_login(login.get(sid))
        risk = 'High' if isinstance(days, int) and days >= (s3_range[0]+3) else 'Medium'
        rows.append([st['last'],st['first'],sid,st['course'],st.get('discipline_class',''),st.get('discipline_teacher',''),st['email'],total,days,last_str,risk,'Re-engage; offer support'])
    write_data_rows(ws, rows, start_row=6); autosize(ws,[22,18,12,10,10,18,38,12,12,14,12,30]); ws.freeze_panes = 'A6'

    # S4
    ws, sids = make_seg_tab('S4',f'Active W{current_week} Absent')
    just_dropped = []; long_silent = []
    if prev_key:
        for sid in sids:
            h = hits[sid]
            if h[prev_key] > 0 and h[curr_key] == 0: just_dropped.append(sid)
            else: long_silent.append(sid)
    else: long_silent = list(sids)
    write_tab_header(ws,f'S4 — Active Then Absent in W{current_week}',f'{len(sids)} total  •  Just Dropped: {len(just_dropped)}  •  Long Silent: {len(long_silent)}',seg_descriptions['S4'],current_week+9,'S4')
    headers_s4 = ['Surname','First Name','Student ID','Course','Disc. Class','Disc. Teacher','Email'] + week_labels[:-1] + ['Group','Priority']
    write_col_headers(ws, headers_s4, row=5)
    def priority_for(h, group):
        if group == 'Just Dropped': basis = h[prev_key]
        else:
            basis = 0
            for k in reversed(week_keys[:-2] if prev_key else week_keys[:-1]):
                if h[k] > 0: basis = h[k]; break
        if basis >= 20: return 'High'
        if basis >= 8: return 'Medium'
        return 'Standard'
    def rows_for_group(sid_list, group):
        out = []
        for sid in sid_list:
            st = students[sid]; h = hits[sid]
            row = [st['last'],st['first'],sid,st['course'],st.get('discipline_class',''),st.get('discipline_teacher',''),st['email']]
            row += [h[k] for k in week_keys[:-1]] + [group, priority_for(h, group)]
            out.append(row)
        pri_order = {'High':0,'Medium':1,'Standard':2}
        def sort_key(r):
            h = hits[r[2]]
            basis = h[prev_key] if (group == 'Just Dropped' and prev_key) else max((h[k] for k in (week_keys[:-2] if prev_key else week_keys[:-1])), default=0)
            return (pri_order[r[-1]], -basis)
        out.sort(key=sort_key); return out
    jd_rows = rows_for_group(just_dropped,'Just Dropped'); ls_rows = rows_for_group(long_silent,'Long Silent')
    n_cols_s4 = len(headers_s4); current_row = 6
    if jd_rows:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=n_cols_s4)
        c = ws.cell(current_row, 1, f'▼ JUST DROPPED — W{current_week-1} activity then silent in W{current_week} ({len(jd_rows)} students). {"Freshest drop signal but distorted by partial week." if is_partial else "Freshest drop signal."}')
        c.font = Font(name='Arial',size=10,bold=True,color=WHITE); c.fill = PatternFill('solid',start_color=RED); c.alignment = Alignment(horizontal='left',vertical='center',indent=1)
        ws.row_dimensions[current_row].height = 22; current_row += 1
        write_data_rows(ws, jd_rows, start_row=current_row); current_row += len(jd_rows)
    if ls_rows:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=n_cols_s4)
        c = ws.cell(current_row, 1, f'▼ LONG SILENT — Active in earlier weeks but zero in W{current_week-1} and W{current_week} ({len(ls_rows)} students). Sustained disengagement; prioritise.')
        c.font = Font(name='Arial',size=10,bold=True,color=WHITE); c.fill = PatternFill('solid',start_color=PURPLE); c.alignment = Alignment(horizontal='left',vertical='center',indent=1)
        ws.row_dimensions[current_row].height = 22; current_row += 1
        write_data_rows(ws, ls_rows, start_row=current_row)
    autosize(ws,[22,18,12,10,10,18,38]+[8]*(current_week-1)+[14,12]); ws.freeze_panes = 'A6'

    # S5
    ws, sids = make_seg_tab('S5','Late Arrivals')
    write_tab_header(ws,f'S5 — Late Arrivals + W{current_week-1} Returners' if prev_key else 'S5 — Late Arrivals',
        f'{len(sids)} students with zero W{current_week-1} hits but appearing in W{current_week}' if prev_key else f'{len(sids)} students',seg_descriptions['S5'],12,'S5')
    write_col_headers(ws,['Surname','First Name','Student ID','Course','Disc. Class','Disc. Teacher','Email','Earlier Total',f'W{current_week} Hits',f'W{current_week} Daily Avg','Type','Watch Status'],row=5)
    rows = []
    for sid in sids:
        st = students[sid]; h = hits[sid]
        prior = sum(h[k] for k in week_keys[:-2]) if current_week >= 2 else 0
        kind = 'True late arrival' if prior == 0 else f'W{current_week-1} absentee returned'
        avg = h[curr_key] / curr_days if curr_days > 0 else 0
        ws_status = 'Promising' if h[curr_key] >= 15 else ('Monitor' if h[curr_key] >= 5 else 'At Risk')
        rows.append([st['last'],st['first'],sid,st['course'],st.get('discipline_class',''),st.get('discipline_teacher',''),st['email'],prior,h[curr_key],round(avg,1),kind,ws_status])
    rows.sort(key=lambda r: -r[8]); write_data_rows(ws, rows, start_row=6)
    autosize(ws,[22,18,12,10,10,18,38,12,10,12,22,14]); ws.freeze_panes = 'A6'

    # S6
    ws, sids = make_seg_tab('S6','Fading Engagers')
    write_tab_header(ws,'S6 — Fading Engagers',f'{len(sids)} students whose daily-average dropped 50%+ from W{current_week-1} to W{current_week}',seg_descriptions['S6'],current_week+9,'S6')
    headers_s6 = ['Surname','First Name','Student ID','Course','Disc. Class','Disc. Teacher','Email'] + week_labels + [f'W{current_week-1} Daily',f'W{current_week} Daily']
    write_col_headers(ws, headers_s6, row=5)
    rows = []
    for sid in sids:
        st = students[sid]; h = hits[sid]
        row = [st['last'],st['first'],sid,st['course'],st.get('discipline_class',''),st.get('discipline_teacher',''),st['email']]
        row += [h[k] for k in week_keys] + [round(h[prev_key]/prev_days,1), round(h[curr_key]/curr_days,1)]
        rows.append(row)
    rows.sort(key=lambda r: -(r[-2]-r[-1])); write_data_rows(ws, rows, start_row=6)
    autosize(ws,[22,18,12,10,10,18,38]+[8]*current_week+[11,11]); ws.freeze_panes = 'A6'

    # S7
    ws, sids = make_seg_tab('S7','Sustained Participants')
    write_tab_header(ws,'S7 — Sustained Participants',
        (f'{len(sids)} students maintaining engagement through W{current_week} (UNDERSTATED — partial week)' if is_partial else f'{len(sids)} students maintaining engagement through W{current_week}'),
        seg_descriptions['S7'],current_week+10,'S7')
    headers_s7 = ['Surname','First Name','Student ID','Course','Disc. Class','Disc. Teacher','Email'] + week_labels + [f'W{current_week-1} Daily',f'W{current_week} Daily','Trend']
    write_col_headers(ws, headers_s7, row=5)
    rows = []
    for sid in sids:
        st = students[sid]; h = hits[sid]
        prev_avg = h[prev_key]/prev_days; curr_avg = h[curr_key]/curr_days
        trend = 'Growing' if curr_avg > prev_avg*1.1 else ('Stable' if curr_avg >= prev_avg*0.85 else 'Slight Dip')
        row = [st['last'],st['first'],sid,st['course'],st.get('discipline_class',''),st.get('discipline_teacher',''),st['email']]
        row += [h[k] for k in week_keys] + [round(prev_avg,1), round(curr_avg,1), trend]
        rows.append(row)
    rows.sort(key=lambda r: -r[-2]); write_data_rows(ws, rows, start_row=6)
    autosize(ws,[22,18,12,10,10,18,38]+[8]*current_week+[11,11,12]); ws.freeze_panes = 'A6'

    # Program Leaderboard
    ws = wb.create_sheet('Program Leaderboard')
    write_tab_header(ws,'Program Leaderboard','Ranked by hits per enrolled (weighted engagement)','Programs ordered by total hits divided by enrolled count. Gold/silver/bronze top 3.',6+current_week)
    headers_pl = ['Rank','Course','Enrolled','Active','Active Rate %','Hits / Enrolled','Total Hits'] + week_labels
    write_col_headers(ws, headers_pl, row=5)
    prog = {}
    for sid, st in students.items():
        course = st['course'] or 'UNKNOWN'; h = hits[sid]
        if course not in prog: prog[course] = {'enr':0,'act':0,'tot':0,**{k:0 for k in week_keys}}
        prog[course]['enr'] += 1; total = sum(h[k] for k in week_keys)
        if total > 0: prog[course]['act'] += 1
        prog[course]['tot'] += total
        for k in week_keys: prog[course][k] += h[k]
    plist = sorted(prog.items(), key=lambda kv: -(kv[1]['tot']/max(kv[1]['enr'],1)))
    rows = []
    for i, (course, p) in enumerate(plist, 1):
        row = [i,course,p['enr'],p['act'],round(100*p['act']/max(p['enr'],1),1),round(p['tot']/max(p['enr'],1),1),p['tot']] + [p[k] for k in week_keys]
        rows.append(row)
    write_data_rows(ws, rows, start_row=6)
    for ri in range(6,6+len(rows)): ws.cell(ri,6).font = Font(name='Arial',size=10,bold=True)
    medals = ['FFD700','C0C0C0','CD7F32']
    for i in range(min(3,len(rows))):
        for ci in range(1,len(headers_pl)+1): ws.cell(6+i,ci).fill = PatternFill('solid',start_color=medals[i])
    autosize(ws,[6,12,10,10,12,14,12]+[10]*current_week); ws.freeze_panes = 'A6'

    # Top 20
    ws = wb.create_sheet('Top 20 Individual')
    write_tab_header(ws,'Top 20 Individual Engagement',f'Ranked by total hits across W1-W{current_week}','Top 20 students by total hits with per-week breakdown.',8+current_week)
    headers_t20 = ['Rank','Surname','First Name','Student ID','Course','Disc. Class','Disc. Teacher','Total'] + week_labels
    write_col_headers(ws, headers_t20, row=5)
    ranked = sorted(students.keys(), key=lambda s: -sum(hits[s][k] for k in week_keys))[:20]
    rows = []
    for i, sid in enumerate(ranked, 1):
        st = students[sid]; h = hits[sid]
        row = [i,st['last'],st['first'],sid,st['course'],st.get('discipline_class',''),st.get('discipline_teacher',''),sum(h[k] for k in week_keys)] + [h[k] for k in week_keys]
        rows.append(row)
    write_data_rows(ws, rows, start_row=6)
    autosize(ws,[6,22,18,12,10,10,18,10]+[8]*current_week); ws.freeze_panes = 'A6'
    return wb, counts

# ===========================================================================
# SHARED HELPER FOR GROUP-BASED REPORTS (Program & Class)
# ===========================================================================

# ===========================================================================
# SHARED HELPER FOR GROUP-BASED REPORTS (Program & Class)
# ===========================================================================
def _build_grouped_report(wb, title_prefix, groups, group_label_col, students, login, hits, seg,
                          current_week, prev_days, curr_days, is_partial, latest_date, login_window,
                          gc_data=None, gc_labels=None, extra_summary_cols=None):
    week_keys = [f'w{i}' for i in range(1, current_week + 1)]
    week_labels = [f'W{i}' for i in range(1, current_week + 1)]
    curr_key = f'w{current_week}'
    prev_key = f'w{current_week - 1}' if current_week > 1 else None
    seg_codes = [f'S{i}' for i in range(1, 8)]
    if extra_summary_cols is None:
        extra_summary_cols = []

    gc_active = gc_data is not None and gc_labels is not None and len(gc_labels) > 0

    # Summary cross-tab
    ws = wb.create_sheet('Summary')
    partial_note = ' (PARTIAL)' if is_partial else ''
    write_tab_header(ws,
        f'{title_prefix} \u2014 W{current_week}{partial_note}',
        f'Enrolled {len(students)}  \u2022  {len(groups)} groups  \u2022  Latest data: {latest_date.strftime("%b %-d %Y")}',
        f'Segment counts per {group_label_col.lower()}. One sheet per group follows.',
        len(seg_codes) + len(extra_summary_cols) + 3)

    headers = [group_label_col] + [name for name, _ in extra_summary_cols] + ['Enrolled'] + seg_codes + ['Active Rate %']
    write_col_headers(ws, headers, row=5)

    group_order = sorted(groups.keys(), key=lambda k: (-len(groups[k]['sids']), str(k)))
    rows = []
    extra_count = len(extra_summary_cols)
    for gk in group_order:
        info = groups[gk]; sids = info['sids']; enrolled = len(sids)
        row = [gk] + [fn(info) for _, fn in extra_summary_cols] + [enrolled]
        active = 0
        for sc in seg_codes:
            n = sum(1 for sid in sids if seg.get(sid) == sc)
            row.append(n)
            if sc not in ('S1', 'S2'): active += n
        row.append(round(100 * active / max(enrolled, 1), 1))
        rows.append(row)

    totals = ['TOTAL'] + ['' for _ in extra_summary_cols] + [len(students)]
    for i, sc in enumerate(seg_codes):
        totals.append(sum(r[i + 2 + extra_count] for r in rows))
    total_active = sum(r[-1] * r[1 + extra_count] / 100 for r in rows)
    totals.append(round(100 * total_active / max(len(students), 1), 1))
    rows.append(totals)
    write_data_rows(ws, rows, start_row=6)

    seg_start_col = 2 + extra_count + 1
    for ci, sc in enumerate(seg_codes):
        col = seg_start_col + ci
        fill_colour = SEG_COLOURS.get(sc)
        if fill_colour:
            ws.cell(5, col).fill = PatternFill('solid', start_color=fill_colour)
            ws.cell(5, col).font = Font(name='Arial', size=10, bold=True, color=WHITE)

    totals_row = 6 + len(rows) - 1
    for ci in range(1, len(headers) + 1):
        ws.cell(totals_row, ci).font = Font(name='Arial', size=10, bold=True)
        ws.cell(totals_row, ci).fill = PatternFill('solid', start_color=LIGHT)
        ws.cell(totals_row, ci).border = thin_border()

    widths = [28] + [24] * extra_count + [10] + [8] * len(seg_codes) + [14]
    autosize(ws, widths)
    ws.freeze_panes = 'A6'

    # Segment Legend
    legend_start = totals_row + 2
    legend_span = max(3 + extra_count, 3)
    ws.merge_cells(start_row=legend_start, start_column=1, end_row=legend_start, end_column=legend_span)
    c = ws.cell(legend_start, 1, 'Segment Legend')
    c.font = Font(name='Arial', size=11, bold=True, color=WHITE)
    c.fill = PatternFill('solid', start_color=NAVY)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[legend_start].height = 22
    seg_legend = [
        ('S1', 'Never Engaged'),
        ('S2', 'Pre-Teaching Ghosts'),
        ('S3', 'W1 Early Drop-Offs'),
        ('S4', f'Active then W{current_week} Absent'),
        ('S5', 'Late Arrivals + Returners'),
        ('S6', 'Fading Engagers'),
        ('S7', 'Sustained Participants'),
    ]
    for li, (code, label) in enumerate(seg_legend):
        row_i = legend_start + 1 + li
        c_code = ws.cell(row_i, 1, code)
        c_code.font = Font(name='Arial', size=10, bold=True, color=WHITE)
        c_code.fill = PatternFill('solid', start_color=SEG_COLOURS.get(code, ACCENT))
        c_code.alignment = Alignment(horizontal='center', vertical='center')
        c_code.border = thin_border()
        ws.merge_cells(start_row=row_i, start_column=2, end_row=row_i, end_column=legend_span)
        c_label = ws.cell(row_i, 2, label)
        c_label.font = Font(name='Arial', size=10)
        c_label.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c_label.border = thin_border()
        for extra_ci in range(3, legend_span + 1):
            ws.cell(row_i, extra_ci).border = thin_border()

    next_row = legend_start + 1 + len(seg_legend) + 1

    # Four rate tables (if GC data present)
    if gc_active:
        n_summary_cols = len(seg_codes) + extra_count + 3
        next_row = _write_all_rate_tables(
            ws, next_row, gc_data, gc_labels, students, seg, n_summary_cols
        )

    # Assessment Detail tab
    if gc_active:
        _write_assessment_detail_sheet(
            wb, gc_data, gc_labels, groups, group_label_col, current_week, is_partial
        )

    group_sheet_names = {}

    # Per-group sheets
    seg_order = {f'S{i}': i for i in range(1, 8)}
    base_headers = (
        ['Segment', 'Surname', 'First Name', 'Student ID', 'Course',
         'Disc. Subject', 'Disc. Class', 'Disc. Teacher', 'Email']
        + week_labels + ['Total Hits']
    )
    if prev_key:
        base_headers += [f'W{current_week-1} Daily', f'W{current_week} Daily']
    base_headers += ['Total Logins', 'Last Login', 'Days Since']
    if gc_active:
        base_headers += gc_labels

    for gk in group_order:
        info = groups[gk]; sids = info['sids']
        sheet_name = str(gk)[:31] or 'Unknown'
        existing_sheets = [s.title for s in wb.worksheets]
        if sheet_name in existing_sheets:
            sheet_name = (sheet_name[:28] + '_2')[:31]
        ws_g = wb.create_sheet(sheet_name)
        group_sheet_names[gk] = sheet_name
        seg_counts_str = ', '.join(
            f'{sc}: {sum(1 for sid in sids if seg.get(sid) == sc)}' for sc in seg_codes
        )
        subtitle = seg_counts_str
        teacher = info.get('teacher', '')
        if teacher:
            subtitle += f'  \u2022  Teacher: {teacher}'
        write_tab_header(
            ws_g, f'{gk} \u2014 {len(sids)} students', subtitle,
            f'All students sorted by segment then surname. W{current_week}{"(partial)" if is_partial else ""}.',
            len(base_headers)
        )
        write_col_headers(ws_g, base_headers, row=5)

        sorted_sids = sorted(
            sids,
            key=lambda sid: (
                seg_order.get(seg.get(sid, 'S1'), 99),
                students[sid]['last'].lower(),
                students[sid]['first'].lower(),
            )
        )
        rows_g = []
        for sid in sorted_sids:
            st = students[sid]; h = hits[sid]; l = login.get(sid)
            days_since, last_str, total_logins = fmt_login(l)
            total_hits = sum(h[k] for k in week_keys)
            row = [
                seg.get(sid, '?'), st['last'], st['first'], sid, st['course'],
                st.get('discipline_subject', ''), st.get('discipline_class', ''),
                st.get('discipline_teacher', ''), st['email'],
            ]
            row += [h[k] for k in week_keys] + [total_hits]
            if prev_key:
                row += [
                    round(h[prev_key] / prev_days, 1) if prev_days else 0,
                    round(h[curr_key] / curr_days, 1) if curr_days else 0,
                ]
            row += [total_logins, last_str, days_since]
            if gc_active:
                for label in gc_labels:
                    entry = gc_data.get(sid, {}).get(label, {})
                    row.append(gc_display_value(entry))
            rows_g.append(row)

        write_data_rows(ws_g, rows_g, start_row=6)

        # Segment colour on col 1
        for ri, sid in enumerate(sorted_sids):
            fill_colour = SEG_COLOURS.get(seg.get(sid, ''))
            if fill_colour:
                ws_g.cell(6 + ri, 1).fill = PatternFill('solid', start_color=fill_colour)
                ws_g.cell(6 + ri, 1).font = Font(name='Arial', size=10, bold=True, color=WHITE)

        # Assessment cell colouring based on final outcome
        if gc_active:
            no_sub_fill        = PatternFill('solid', start_color='FADBD8')
            no_sub_font        = Font(name='Arial', size=10, color=RED, bold=True)
            unsat_fill         = PatternFill('solid', start_color='FDEBD0')
            unsat_font         = Font(name='Arial', size=10, color=ORANGE, bold=True)
            needs_grading_fill = PatternFill('solid', start_color='FEF9E7')
            needs_grading_font = Font(name='Arial', size=10, color='7D6608')
            sat_fill           = PatternFill('solid', start_color='D5F5E3')
            sat_font           = Font(name='Arial', size=10, color='1A7A3A', bold=True)

            gc_start_col = len(base_headers) - len(gc_labels) + 1
            for ri, sid in enumerate(sorted_sids):
                for ci_offset, label in enumerate(gc_labels):
                    cell = ws_g.cell(6 + ri, gc_start_col + ci_offset)
                    entry = gc_data.get(sid, {}).get(label, {})
                    outcome = gc_final_outcome(entry)
                    if outcome == 'No Submission':
                        cell.fill = no_sub_fill;        cell.font = no_sub_font
                    elif outcome == 'Unsatisfactory':
                        cell.fill = unsat_fill;         cell.font = unsat_font
                    elif outcome == 'Needs Grading':
                        cell.fill = needs_grading_fill; cell.font = needs_grading_font
                    else:  # Satisfactory (original or via resubmit)
                        cell.fill = sat_fill;           cell.font = sat_font

        widths = [8, 22, 18, 12, 10, 24, 10, 18, 38] + [8] * current_week + [10]
        if prev_key:
            widths += [11, 11]
        widths += [12, 14, 12]
        if gc_active:
            widths += [22] * len(gc_labels)
        autosize(ws_g, widths)
        ws_g.freeze_panes = 'A6'

    # Hyperlinks from summary to group tabs
    for i, gk in enumerate(group_order):
        sn = group_sheet_names.get(gk)
        if sn:
            cell = ws.cell(6 + i, 1)
            cell.hyperlink = f"#'{sn}'!A1"
            cell.font = Font(name='Arial', size=10, color='2980B9', underline='single')

    # Class Index tab
    _write_class_index_sheet(
        wb, title_prefix, groups, group_label_col, students, seg,
        gc_data if gc_active else None,
        gc_labels if gc_active else None,
        group_sheet_names, current_week, is_partial,
    )

    # Move Class Index to position 2 (after Summary, before Assessment Detail)
    target_idx = 2
    current_idx = next(
        (i for i, s in enumerate(wb._sheets) if s.title == 'Class Index'), None
    )
    if current_idx is not None and current_idx != target_idx:
        sheet = wb._sheets.pop(current_idx)
        wb._sheets.insert(target_idx, sheet)

    return wb


def _write_program_leaderboard_sheet(wb, title_prefix, groups, students, seg,
                                      gc_data, gc_labels, group_sheet_names,
                                      current_week, is_partial):
    gc_active = gc_data is not None and gc_labels is not None and len(gc_labels) > 0
    partial_note = ' (PARTIAL)' if is_partial else ''

    n_fixed = 5
    n_assess_cols = (len(gc_labels) * 2) if gc_active else 0
    n_cols = n_fixed + n_assess_cols

    ws = wb.create_sheet('Program Leaderboard')

    write_tab_header(
        ws,
        f'{title_prefix} \u2014 Program Leaderboard{partial_note}',
        f'{len(groups)} programs  \u2022  Active = S5+S6+S7  \u2022  At Risk = S4',
        ('Sub = submitted/enrolled.  Pass = passed (incl. resubmit)/enrolled.  '
         'Green = 100%, Orange >= 50%, Red < 50%.'
         if gc_active else
         'Active = S5+S6+S7.  At Risk = S4.'),
        n_cols,
    )

    col_headers = ['#', 'Program', 'Enrolled', 'Active', 'At Risk']
    if gc_active:
        for lbl in gc_labels:
            col_headers += [f'{lbl} Sub', f'{lbl} Pass']

    for ci, h in enumerate(col_headers, 1):
        c = ws.cell(5, ci, h)
        c.font = Font(name='Arial', size=10, bold=True, color=WHITE)
        c.fill = PatternFill('solid', start_color=NAVY)
        c.alignment = Alignment(
            horizontal='left' if ci == 2 else 'center', vertical='center', wrap_text=True
        )
        c.border = thin_border()
    ws.row_dimensions[5].height = 30

    fill_green  = PatternFill('solid', start_color='D5F5E3')
    fill_orange = PatternFill('solid', start_color='FDEBD0')
    fill_red    = PatternFill('solid', start_color='FADBD8')
    font_green  = Font(name='Arial', size=10, color='1A7A3A', bold=True)
    font_orange = Font(name='Arial', size=10, color=ORANGE, bold=True)
    font_red    = Font(name='Arial', size=10, color=RED, bold=True)

    def rate_style(cell, n, denom):
        rate = n / denom if denom else 0
        cell.value = f'{n}/{denom}'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border()
        if rate == 1.0:
            cell.fill = fill_green;  cell.font = font_green
        elif rate >= 0.5:
            cell.fill = fill_orange; cell.font = font_orange
        else:
            cell.fill = fill_red;   cell.font = font_red

    group_order = sorted(groups.keys(), key=lambda k: (-len(groups[k]['sids']), str(k)))

    for ri, gk in enumerate(group_order):
        excel_row = 6 + ri
        sids = groups[gk]['sids']
        enrolled = len(sids)
        active   = sum(1 for sid in sids if seg.get(sid) in ('S5', 'S6', 'S7'))
        at_risk  = sum(1 for sid in sids if seg.get(sid) == 'S4')

        row_fill = PatternFill('solid', start_color=ALT_ROW) if ri % 2 == 0 else None

        fixed_vals = [ri + 1, gk, enrolled, active, at_risk]
        for ci, val in enumerate(fixed_vals, 1):
            c = ws.cell(excel_row, ci, val)
            c.font = Font(name='Arial', size=10)
            if row_fill:
                c.fill = row_fill
            c.alignment = Alignment(
                horizontal='left' if ci == 2 else 'center', vertical='center'
            )
            c.border = thin_border()

        if at_risk > 0:
            ws.cell(excel_row, 5).font = Font(name='Arial', size=10, color=RED, bold=True)

        sn = group_sheet_names.get(gk)
        if sn:
            cell = ws.cell(excel_row, 2)
            cell.hyperlink = f"#'{sn}'!A1"
            cell.font = Font(name='Arial', size=10, color='2980B9', underline='single')

        if gc_active:
            for ai, label in enumerate(gc_labels):
                sub_col  = n_fixed + 1 + ai * 2
                pass_col = n_fixed + 2 + ai * 2
                n_sub  = sum(
                    1 for sid in sids
                    if gc_submitted(gc_data.get(sid, {}).get(label, {}))
                )
                n_pass = sum(
                    1 for sid in sids
                    if gc_passed(gc_data.get(sid, {}).get(label, {}))
                )
                rate_style(ws.cell(excel_row, sub_col),  n_sub,  enrolled)
                rate_style(ws.cell(excel_row, pass_col), n_pass, enrolled)

    totals_row = 6 + len(group_order)
    total_enrolled = sum(len(groups[gk]['sids']) for gk in group_order)
    total_active   = sum(
        sum(1 for sid in groups[gk]['sids'] if seg.get(sid) in ('S5', 'S6', 'S7'))
        for gk in group_order
    )
    total_at_risk  = sum(
        sum(1 for sid in groups[gk]['sids'] if seg.get(sid) == 'S4')
        for gk in group_order
    )
    total_fixed = ['', 'TOTAL', total_enrolled, total_active, total_at_risk]
    for ci, val in enumerate(total_fixed, 1):
        c = ws.cell(totals_row, ci, val)
        c.font = Font(name='Arial', size=10, bold=True)
        c.fill = PatternFill('solid', start_color=LIGHT)
        c.border = thin_border()
        c.alignment = Alignment(horizontal='left' if ci == 2 else 'center', vertical='center')

    if gc_active:
        for ai, label in enumerate(gc_labels):
            sub_col  = n_fixed + 1 + ai * 2
            pass_col = n_fixed + 2 + ai * 2
            t_sub  = sum(
                1 for gk in group_order for sid in groups[gk]['sids']
                if gc_submitted(gc_data.get(sid, {}).get(label, {}))
            )
            t_pass = sum(
                1 for gk in group_order for sid in groups[gk]['sids']
                if gc_passed(gc_data.get(sid, {}).get(label, {}))
            )
            for col, n in ((sub_col, t_sub), (pass_col, t_pass)):
                c = ws.cell(totals_row, col, f'{n}/{total_enrolled}')
                c.font = Font(name='Arial', size=10, bold=True)
                c.fill = PatternFill('solid', start_color=LIGHT)
                c.border = thin_border()
                c.alignment = Alignment(horizontal='center', vertical='center')

    widths = [4, 20, 10, 10, 10] + ([9, 9] * len(gc_labels) if gc_active else [])
    autosize(ws, widths)
    ws.freeze_panes = 'C6'
    return ws


def build_program_workbook(subject_code, students, login, hits, seg, current_week,
                           prev_days, curr_days, is_partial, latest_date, login_window,
                           gc_data=None, gc_labels=None):
    wb = Workbook(); wb.remove(wb.active)
    programs = {}
    for sid, st in students.items():
        cc = st.get('course_code') or st.get('course') or 'UNKNOWN'
        programs.setdefault(cc, {'sids': []})
        programs[cc]['sids'].append(sid)
    title_prefix = f'{subject_code} Program Report'
    wb = _build_grouped_report(
        wb, title_prefix, programs, 'Program',
        students, login, hits, seg, current_week, prev_days, curr_days,
        is_partial, latest_date, login_window,
        gc_data=gc_data, gc_labels=gc_labels,
    )

    reserved = {'Summary', 'Assessment Detail', 'Class Index'}
    group_sheet_names = {}
    for gk in programs:
        sheet_name = str(gk)[:31] or 'Unknown'
        if sheet_name in [s.title for s in wb.worksheets] and sheet_name not in reserved:
            group_sheet_names[gk] = sheet_name

    gc_active = gc_data is not None and gc_labels is not None and len(gc_labels) > 0
    _write_program_leaderboard_sheet(
        wb, title_prefix, programs, students, seg,
        gc_data if gc_active else None,
        gc_labels if gc_active else None,
        group_sheet_names, current_week, is_partial,
    )

    current_idx = next(
        (i for i, s in enumerate(wb._sheets) if s.title == 'Program Leaderboard'), None
    )
    if current_idx is not None and current_idx != 1:
        sheet = wb._sheets.pop(current_idx)
        wb._sheets.insert(1, sheet)

    return wb


def build_class_workbook(subject_code, students, login, hits, seg, current_week,
                         prev_days, curr_days, is_partial, latest_date, login_window,
                         gc_data=None, gc_labels=None):
    wb = Workbook(); wb.remove(wb.active)
    classes = {}
    for sid, st in students.items():
        ds = st.get('discipline_subject', '') or ''
        dc = st.get('discipline_class', '') or ''
        dt = st.get('discipline_teacher', '') or ''
        key = f'{ds} \u2014 {dc}'.strip(' \u2014') if ds else (dc if dc else 'Unassigned')
        if key not in classes:
            classes[key] = {'sids': [], 'teacher': dt}
        classes[key]['sids'].append(sid)
        if dt and not classes[key]['teacher']:
            classes[key]['teacher'] = dt
    return _build_grouped_report(
        wb, f'{subject_code} Class Report', classes, 'Class',
        students, login, hits, seg, current_week, prev_days, curr_days,
        is_partial, latest_date, login_window,
        gc_data=gc_data, gc_labels=gc_labels,
        extra_summary_cols=[('Teacher', lambda info: info.get('teacher', ''))],
    )


# ===========================================================================
# STREAMLIT UI
# ===========================================================================
st.set_page_config(page_title='WSUTC Engagement Report', layout='wide', page_icon='\U0001f4ca')
st.title('WSUTC Student Engagement Report')
st.caption('Upload Blackboard exports for one subject. Generates an Excel workbook with engagement segmentation.')

with st.expander('Instructions', expanded=False):
    st.markdown(
        'Upload the **class list** (.xls or .xlsx), the **login report** (.xlsx), '
        'and **all relevant usage report files** (.xls).  \n'
        '**Grade Centre (.xls, optional):** Four rate tables are shown per report: '
        'Submission, Pass, Resubmission, and Resubmission Pass rates.  \n'
        'Per-student assessment cells show final outcome (green = passed, including via resubmit).  \n'
        'Usage files can overlap; the most recent data wins.  \n'
        'The current teaching week is auto-detected from the latest day with data.'
    )

col1, col2 = st.columns(2)
with col1:
    classlist_file = st.file_uploader('Class list (.xls / .xlsx)', type=['xls', 'xlsx'], key='cl')
    login_file = st.file_uploader('Login report (.xlsx)', type=['xlsx'], key='lr')
with col2:
    usage_files = st.file_uploader(
        'Usage report files (.xls) \u2014 upload all that apply',
        type=['xls'], accept_multiple_files=True, key='uf',
    )
    gc_file = st.file_uploader('Grade Centre (.xls) \u2014 optional', type=['xls'], key='gc')

run_btn = st.button(
    'Generate report', type='primary',
    disabled=not (classlist_file and login_file and usage_files),
)

if run_btn:
    try:
        with st.spinner('Loading class list...'):
            subject_code, students = load_classlist(classlist_file.getvalue())
        st.success(f'**{subject_code}** \u2022 {len(students)} enrolled (after exclusions)')

        with st.spinner('Loading login report...'):
            login, win_start, win_end = load_login_report(login_file.getvalue())
        if win_start and win_end:
            login_window_str = f'{win_start.strftime("%b %-d")} - {win_end.strftime("%b %-d %Y")}'
            st.info(f'Login window detected: **{login_window_str}**')
        else:
            login_window_str = 'unknown (could not parse)'
            st.warning('Could not auto-detect login window from report. Check the file format.')

        with st.spinner(f'Parsing {len(usage_files)} usage file(s)...'):
            merged = merge_usage_files([f.getvalue() for f in usage_files])
        st.write(f'Parsed {len(merged)} unique day-records from usage files.')

        cap_latest = win_end if win_end else None
        current_week, days_in, latest = detect_current_week(merged, override_latest=cap_latest)
        if current_week is None:
            st.error('No usage data found on or after Mar 2 2026. Cannot determine current week.')
            st.stop()
        is_partial = days_in < 7
        curr_days = days_in
        prev_days = 7
        if current_week == 1:
            prev_days = curr_days

        partial_msg = f' (PARTIAL \u2014 {curr_days} days)' if is_partial else ' (full)'
        st.success(
            f'Detected current week: **W{current_week}**{partial_msg}  \u2022  '
            f'Latest data: **{latest.strftime("%b %-d %Y")}**'
        )
        if current_week == 1:
            st.warning('W1 is the only week with data. S5/S6/S7 will be empty (no comparison week available).')

        with st.spinner('Computing weekly hits...'):
            hits = bucket_by_week(merged, students, current_week, max_date=latest)

        with st.spinner('Classifying students...'):
            seg, s2_thresh, s3_rng = classify(
                students, login, hits, current_week, prev_days, curr_days
            )

        counts = {f'S{i}': 0 for i in range(1, 8)}
        for s in seg.values():
            counts[s] = counts.get(s, 0) + 1
        assert sum(counts.values()) == len(students), \
            f'Segment sum {sum(counts.values())} != enrolled {len(students)}'

        st.subheader('Segment counts')
        cols = st.columns(7)
        for i, (code, _) in enumerate([
            ('S1', 'Never'), ('S2', 'Ghosts'), ('S3', 'W1 drop'),
            ('S4', f'W{current_week} absent'), ('S5', 'Late/return'),
            ('S6', 'Fading'), ('S7', 'Sustained'),
        ]):
            cols[i].metric(code, counts[code], f'{counts[code]/len(students)*100:.1f}%')

        st.subheader('Standing checks')
        missing = sum(1 for s in students if s not in login)
        st.write(f'\u2022 Enrolled: **{len(students)}**')
        st.write(f'\u2022 Missing from login report: **{missing}**')
        st.write(f'\u2022 S2 days-since threshold: **>= {s2_thresh}**')
        st.write(f'\u2022 S3 days-since range: **{s3_rng[0]}-{s3_rng[1]}**')
        if current_week > 1:
            st.write(f'\u2022 Comparison: W{current_week-1} (7d) vs W{current_week} ({curr_days}d)')
        else:
            st.write('\u2022 No prior week to compare')

        with st.spinner('Building workbook...'):
            wb, _ = build_workbook(
                subject_code, students, login, hits, seg,
                current_week, prev_days, curr_days, is_partial, latest,
                login_window_str, s2_thresh, s3_rng,
            )
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)

        date_str = latest.strftime('%Y%m%d')
        suffix = '_Partial' if is_partial else ''
        filename = f'{subject_code}_Engagement_Report_W{current_week}_{date_str}{suffix}.xlsx'
        st.download_button(
            '\u2b07 Download workbook', data=buf, file_name=filename,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            type='primary',
        )

        gc_data = None
        gc_labels = None
        if gc_file is not None:
            with st.spinner('Parsing grade centre...'):
                gc_data, gc_labels = load_grade_centre(gc_file.getvalue())
            matched = sum(1 for sid in students if sid in gc_data)
            st.info(
                f'Grade Centre: detected **{len(gc_labels)}** assessments '
                f'({", ".join(gc_labels)}) \u2022 matched **{matched}/{len(students)}** students'
            )

        has_programs = any(st_data.get('course_code') for st_data in students.values())
        if has_programs:
            with st.spinner('Building program report...'):
                wb_prog = build_program_workbook(
                    subject_code, students, login, hits, seg,
                    current_week, prev_days, curr_days, is_partial, latest,
                    login_window_str, gc_data=gc_data, gc_labels=gc_labels,
                )
                buf_prog = io.BytesIO(); wb_prog.save(buf_prog); buf_prog.seek(0)
            prog_filename = f'{subject_code}_Program_Report_W{current_week}_{date_str}{suffix}.xlsx'
            st.download_button(
                '\u2b07 Download program report', data=buf_prog, file_name=prog_filename,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                key='prog_dl',
            )

            has_classes = any(st_data.get('discipline_class') for st_data in students.values())
            if has_classes:
                with st.spinner('Building class report...'):
                    wb_class = build_class_workbook(
                        subject_code, students, login, hits, seg,
                        current_week, prev_days, curr_days, is_partial, latest,
                        login_window_str, gc_data=gc_data, gc_labels=gc_labels,
                    )
                    buf_class = io.BytesIO(); wb_class.save(buf_class); buf_class.seek(0)
                class_filename = f'{subject_code}_Class_Report_W{current_week}_{date_str}{suffix}.xlsx'
                st.download_button(
                    '\u2b07 Download class report', data=buf_class, file_name=class_filename,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    key='class_dl',
                )

    except Exception as e:
        st.error(f'Error: {e}')
        import traceback
        st.code(traceback.format_exc())
